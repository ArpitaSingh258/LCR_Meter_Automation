# ============================================================
#  GUI.py  —  CSIR-NPL LCR Automation Software  v5.0
#  Agilent E4980A  |  GPIB0::17::INSTR
#
#  v5: Futuristic dashboard — glassmorphism, gradients,
#      sidebar nav, large live readings, neon graph
#      ALL backend logic 100% preserved from v4
# ============================================================

import json
import os
import sys
import traceback
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import threading
import time
import csv
import os
import math
import statistics as _stats
from datetime import datetime
from collections import deque
import json

import pyvisa
import math

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.ticker as mticker

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image as RLImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# ════════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ════════════════════════════════════════════════════════════
# Backgrounds
BG_APP       = "#0F0F1A"   # deepest bg
BG_SIDEBAR   = "#13131F"   # sidebar
BG_HEADER    = "#13131F"   # header bar
BG_CARD      = "#1A1A2E"   # card base
BG_CARD2     = "#16213E"   # slightly lighter card
BG_INPUT     = "#0F3460"   # input fields
BG_GRAPH     = "#080810"   # graph area
BG_WHITE     = "#FFFFFF"   # white card
BG_LIGHT     = "#F8F9FF"   # light card bg

# Gradients simulated as solid for Tkinter
GRAD_START   = "#5B5FFF"   # blue-violet
GRAD_END     = "#C94BFF"   # purple
GRAD_MID     = "#8A4FFF"   # midpoint used as solid

# Accent colours
CYAN         = "#00D9FF"
GREEN        = "#30D158"
RED          = "#FF453A"
PINK         = "#FF2D75"
AMBER        = "#FFD60A"
ORANGE       = "#FF9F0A"
VIOLET       = "#5B5FFF"
PURPLE       = "#C94BFF"

# Mode palette — pastel card backgrounds (like reference)
MODE_CARD_COLORS = [
    "#E8F4FD", "#EDF9F0", "#FEF3E2", "#F3E8FF",
    "#FFE8EC", "#E8FFFA", "#FFF8E8", "#E8E8FF",
    "#F0FFE8", "#FFE8F0", "#E8FFF8", "#F8E8FF",
    "#FFF0E8", "#E8F8FF", "#FFFFE8", "#E8FFEC",
    "#FFE8E8", "#E8F0FF", "#F8FFE8", "#FFE8FB",
]
MODE_CARD_FG = [
    "#1A6FBF", "#1A8A3A", "#BF7A1A", "#7A1ABF",
    "#BF1A3A", "#1ABF9A", "#BF9A1A", "#1A1ABF",
    "#4ABF1A", "#BF1A7A", "#1ABF6A", "#9A1ABF",
    "#BF5A1A", "#1A7ABF", "#8A8A1A", "#1ABF4A",
    "#BF1A1A", "#1A4ABF", "#6ABF1A", "#BF1AAA",
]

# Graph neon lines
G_CYAN       = "#00D9FF"
G_PINK       = "#FF2D75"
G_GREEN      = "#30D158"
G_AMBER      = "#FFD60A"
G_GRID       = "#1A1A30"
MODE_COLORS  = [G_CYAN, G_GREEN, G_AMBER]

# Text
T_WHITE      = "#FFFFFF"
T_LIGHT      = "#F2F2F7"
T_SECOND     = "#8E8E93"
T_DIM        = "#48484A"
T_DARK       = "#1C1C1E"
T_VIOLET     = VIOLET
T_CYAN       = CYAN

# Typography
SUI          = "Segoe UI"
MONO         = "Consolas"

F_TITLE      = (SUI, 22, "bold")
F_SUBTITLE   = (SUI, 10)
F_SECTION    = (SUI, 11, "bold")
F_LABEL      = (SUI, 9)
F_LABEL_B    = (SUI, 9, "bold")
F_BODY       = (SUI, 10)
F_BODY_B     = (SUI, 10, "bold")
F_VALUE_XL   = (MONO, 38, "bold")
F_VALUE_LG   = (MONO, 22, "bold")
F_VALUE_MD   = (SUI, 15, "bold")
F_VALUE_SM   = (SUI, 12, "bold")
F_BTN_LG     = (SUI, 14, "bold")
F_BTN        = (SUI, 10, "bold")
F_SMALL      = (SUI, 9)
F_TINY       = (SUI, 8)
F_NAV        = (SUI, 9, "bold")
F_MONO_SM    = (MONO, 8)
F_MONO_MD    = (MONO, 11, "bold")

# ════════════════════════════════════════════════════════════
#  CONSTANTS
# ════════════════════════════════════════════════════════════
GPIB_ADDRESS = "GPIB0::17::INSTR"
APP_TITLE    = "LCR Meter Automation"
APP_SUB      = "Agilent E4980A Precision LCR Meter"
GRAPH_POINTS = 60
AUTOSAVE_FILE = "lcr_autosave_session.json"
DATALOGGER_MAX_ROWS = 500

MODES = [
    "Cp-D",  "Cp-Q",  "Cp-G",  "Cp-Rp",
    "Cs-D",  "Cs-Q",  "Cs-Rs",
    "Lp-D",  "Lp-Q",  "Lp-G",  "Lp-Rp", "Lp-Rdc",
    "Ls-D",  "Ls-Q",  "Ls-Rs", "Ls-Rdc",
    "Z-θd",  "Z-θr",
    "Y-θd",  "Y-θr",
    "R-X",   "G-B","Vdc-Idc",
]

# Maps each display-friendly mode name (used throughout the UI) to
# the actual SCPI keyword the E4980A's FUNC:IMP command accepts.
# IMPORTANT: these must be the real instrument keywords — sending a
# display string like "Cp-D" directly (as a previous version of this
# mapping did) gets rejected by the instrument with "E-224: Illegal
# parameter value", silently leaving it in whatever mode it was
# already in instead of switching to the one you selected.
MODE_SCPI = {
    "Cp-D": "CPD",   "Cp-Q": "CPQ",   "Cp-G": "CPG",   "Cp-Rp": "CPRP",
    "Cs-D": "CSD",   "Cs-Q": "CSQ",   "Cs-Rs": "CSRS",
    "Lp-D": "LPD",   "Lp-Q": "LPQ",   "Lp-G": "LPG",
    "Lp-Rp": "LPRP", "Lp-Rdc": "LPRDC",
    "Ls-D": "LSD",   "Ls-Q": "LSQ",   "Ls-Rs": "LSRS", "Ls-Rdc": "LSRDC",
    "Z-θd": "ZTD",   "Z-θr": "ZTR",
    "Y-θd": "YTD",   "Y-θr": "YTR",
    "R-X": "RX",     "G-B": "GB",     "Vdc-Idc": "VDCIDC",
}


def scpi_code(mode: str) -> str:
    """Translate a display-friendly mode name (e.g. 'Cp-D') to the
    canonical SCPI/lookup code (e.g. 'CPD'). Falls back to the
    upper-cased input itself for any mode string that's already a
    canonical code (keeps older call sites/CSVs working)."""
    return MODE_SCPI.get(mode, mode.strip().upper())

from decimal import Decimal, ROUND_HALF_UP


def _round_half_up(x: float, ndigits: int = 0) -> float:
    """Round like the E4980A firmware (half-away-from-zero), not
    like Python's built-in round() (half-to-even / "banker's
    rounding"). The two only disagree exactly on .5 boundaries, but
    when they do, banker's rounding was producing a genuine
    off-by-one last-digit mismatch against the panel even for an
    identical raw value."""
    q = Decimal(1).scaleb(-ndigits) if ndigits > 0 else Decimal(1)
    return float(Decimal(repr(x)).quantize(q, rounding=ROUND_HALF_UP))

# ════════════════════════════════════════════════════════════
#  FORMATTER  (merged in from formatter.py — matches Agilent
#  E4980A front-panel display exactly)
# ════════════════════════════════════════════════════════════
# BUG FIX (confirmed against real front-panel photos, 2026-07-07):
# the previous table (SHORT:5, MED:6, LONG:7) was one significant
# digit short across the board. Cross-checking three live captures
# of the E4980A display at MEAS TIME MED —
#   Cp -228.9150 nF   (7 sig figs)   D  1.504065        (7 sig figs)
#   Cs -6.437719 µF   (7 sig figs)   Rs 375.7247 Ω       (7 sig figs)
#   Z   368.0051 Ω    (7 sig figs)   Θ  577.1725 mrad    (7 sig figs)
# all six independent readings at MED are 7 significant digits, not
# 6. This alone explains "fewer digits than the instrument" (issue
# #3) for every MED-aperture measurement. SHORT/LONG below are
# shifted by the same +1 to stay internally consistent, but — unlike
# MED — they are NOT yet confirmed against a photo. Verify them the
# same way (switch MEAS TIME on the instrument, compare one live
# reading against get_sig_figs()'s output) and adjust the two
# numbers below if needed; nothing else in the formatter depends on
# the exact values.
def get_sig_figs(aperture: str) -> int:
    return {
        "SHORT": 6,
        "MED":   7,
        "LONG":  7,
    }.get(aperture.strip().upper(), 7)


def _apply_prefix(value: float, sig_figs: int, prefixes: list) -> str:
    """Engineering-notation formatter matching the E4980A's own
    auto-ranging display logic.

    BUG FIX — unit rollover: the previous version chose the
    engineering prefix from the *unrounded* value, then rounded
    afterwards. That ordering breaks near every unit boundary: e.g.
    999.99962 nF (sig_figs=6) picks the "nF" prefix because
    999.99962/1e-9 >= 1, then rounds to "1000.00 nF" — six sig figs
    but a value that no longer starts with a leading 1-9 digit in
    that unit, and one the instrument would instead display as
    "1.00000 µF" after its own rounding rolls it over to the next
    prefix. Any C/L/R/Y reading landing within half a step of a
    ×1000 (or ×1 for the Ω/H tables) boundary would previously come
    out with a mismatched unit and a wrong number of digits versus
    the panel — this reproduces both issue #3 (extra/fewer digits)
    and issue #8 (wrong engineering unit). Fix: round first in the
    initially-chosen unit, then re-check whether the rounded value
    has spilled into the next prefix up, and re-render in that unit
    if so.
    """
    abs_val = abs(value)
    chosen_div, chosen_unit = prefixes[-1]
    chosen_idx = len(prefixes) - 1
    for idx, (div, unit) in enumerate(prefixes):
        scaled_check = abs_val / div
        if scaled_check >= 1.0:
            chosen_div  = div
            chosen_unit = unit
            chosen_idx  = idx
            break

    def _render(div, sig_figs_local):
        scaled = value / div
        if scaled == 0.0:
            return 0.0, f"0.{'0' * (sig_figs_local - 1)}"
        mag = math.floor(math.log10(abs(scaled)))
        decimal_places = max(0, sig_figs_local - 1 - int(mag))
        factor = 10 ** (sig_figs_local - 1 - int(mag))
        rounded = _round_half_up(scaled * factor) / factor
        # Rounding can itself bump the magnitude up a decade (e.g.
        # 0.99999999 rounds to 1.0) — recompute decimal_places once
        # against the *rounded* value so the digit count still comes
        # out to exactly sig_figs_local, not one too many.
        if rounded != 0.0:
            new_mag = math.floor(math.log10(abs(rounded)))
            if new_mag != mag:
                decimal_places = max(0, sig_figs_local - 1 - int(new_mag))
        return rounded, f"{rounded:.{decimal_places}f}"

    rounded, text = _render(chosen_div, sig_figs)

    # Rollover check: if rounding pushed the magnitude up to (or
    # past) the next larger prefix's threshold, re-render in that
    # unit instead — this is exactly what the instrument's own
    # auto-ranging display does.
    if chosen_idx > 0 and abs(rounded) >= 1000.0 - 1e-9:
        next_div, next_unit = prefixes[chosen_idx - 1]
        # Only re-render if the *next* prefix step is actually a
        # ×1000 relationship (true for every table in this file);
        # guards against odd/non-decade prefix lists silently
        # mis-firing.
        if chosen_div != 0 and round(next_div / chosen_div) == 1000:
            chosen_div, chosen_unit = next_div, next_unit
            rounded, text = _render(chosen_div, sig_figs)

    return f"{text} {chosen_unit}"


def _format_dimensionless(value: float, sig_figs: int, param_type: str = "Q") -> str:
    if value == 0.0:
        return f"0.{'0' * (sig_figs - 1)}"
    mag = math.floor(math.log10(abs(value)))
    decimal_places = max(0, sig_figs - 1 - int(mag))
    factor = 10 ** (sig_figs - 1 - int(mag))
    rounded = _round_half_up(value * factor) / factor
    return f"{rounded:.{decimal_places}f}"


def format_lcr_value(value: float, param_type: str, aperture: str) -> str:
    sig_figs = get_sig_figs(aperture)
    print("SIG FIGS =", sig_figs)
    p = param_type.strip().upper()

    if p == "C":
        return _apply_prefix(value, sig_figs, [
            (1e-3,  "mF"), (1e-6,  "\u00b5F"),
            (1e-9,  "nF"), (1e-12, "pF"),
        ])
    elif p == "L":
        return _apply_prefix(value, sig_figs, [
            (1.0,   "H"), (1e-3,  "mH"),
            (1e-6,  "\u00b5H"), (1e-9,  "nH"),
        ])
    elif p in ("R", "Z", "X", "RS", "RP"):
        return _apply_prefix(value, sig_figs, [
            (1e6,  "M\u03a9"), (1e3,  "k\u03a9"),
            (1.0,  "\u03a9"), (1e-3, "m\u03a9"),
            (1e-6, "\u00b5\u03a9"), (1e-9, "n\u03a9"),
        ])
    elif p in ("Y", "G", "B"):
        return _apply_prefix(value, sig_figs, [
            (1.0,  "S"), (1e-3, "mS"),
            (1e-6, "\u00b5S"), (1e-9, "nS"),
        ])
    elif p == "VDC":
        return _apply_prefix(value, sig_figs, [
            (1.0,  "V"), (1e-3, "mV"), (1e-6, "\u00b5V"),
        ])
    elif p == "IDC":
        return _apply_prefix(value, sig_figs, [
            (1.0,  "A"), (1e-3, "mA"), (1e-6, "\u00b5A"), (1e-9, "nA"),
        ])
    elif p == "D":
        return f"{value:.6f}"
    elif p == "Q":
        return f"{value:.2f}"
    elif p == "THETA_DEG":
        # Degree-mode theta (ZTD/YTD): the instrument shows this as a
        # plain fixed-point angle, no engineering prefix — bounded to
        # ±180°, so there's no unit to auto-range across.
        if value == 0.0:
            decimal_places = sig_figs - 1
        else:
            mag = math.floor(math.log10(abs(value)))
            decimal_places = max(0, sig_figs - 1 - int(mag))
        return f"{value:.{decimal_places}f} \u00b0"
    elif p == "THETA_RAD":
        # BUG FIX: radian-mode theta (ZTR/YTR) was previously routed
        # through the same fixed-point "° " branch as THETA_DEG. But
        # the instrument reports ZTR/YTR angles in *radians*, engineering-
        # prefixed exactly like an impedance or admittance reading (e.g.
        # the front panel showing "577.1725 mrad", confirmed on a live
        # Z-θr capture) — never with a "°" suffix and never as a bare
        # fixed-point number. Using the degree formatter here produced
        # both a wrong unit label and a wrong digit count (issue #8).
        return _apply_prefix(value, sig_figs, [
            (1.0,  "rad"), (1e-3, "mrad"), (1e-6, "\u00b5rad"),
        ])
    elif p == "THETA":
        # Legacy/unknown-variant fallback — kept only for callers that
        # still pass the old undifferentiated "THETA" type; new code
        # should always use THETA_DEG or THETA_RAD via MODE_PARAM_TYPES.
        if value == 0.0:
            decimal_places = sig_figs - 1
        else:
            mag = math.floor(math.log10(abs(value)))
            decimal_places = max(0, sig_figs - 1 - int(mag))
        return f"{value:.{decimal_places}f} \u00b0"
    else:
        return _format_dimensionless(value, sig_figs)


MODE_PARAM_TYPES = {
    "CPD": ("C", "D"), "CPQ": ("C", "Q"), "CPG": ("C", "G"), "CPRP": ("C", "R"),
    "CSD": ("C", "D"), "CSQ": ("C", "Q"), "CSRS": ("C", "RS"),
    "LPD": ("L", "D"), "LPQ": ("L", "Q"), "LPG": ("L", "G"),
    "LPRDC": ("L", "R"), "LPRD": ("L", "R"),
    "LSD": ("L", "D"), "LSQ": ("L", "Q"), "LSRS": ("L", "RS"),"LSRDC": ("L", "R"),
    "ZTD": ("Z", "THETA_DEG"), "ZTR": ("Z", "THETA_RAD"),
    "ZTHD": ("Z", "THETA_DEG"), "ZTHR": ("Z", "THETA_RAD"),
    "YTD": ("Y", "THETA_DEG"), "YTR": ("Y", "THETA_RAD"),
    "RX": ("R", "X"),
    "GB": ("G", "B"),
    "VDCIDC": ("VDC", "IDC"),
}

MODE_LABELS = {
    "CPD": ("Cp", "D"), "CPQ": ("Cp", "Q"), "CPG": ("Cp", "G"), "CPRP": ("Cp", "Rp"),
    "CSD": ("Cs", "D"), "CSQ": ("Cs", "Q"), "CSRS": ("Cs", "Rs"),
    "LPD": ("Lp", "D"), "LPQ": ("Lp", "Q"), "LPG": ("Lp", "G"), "LPRP": ("Lp", "Rp"),
    "LPRDC": ("Lp", "Rdc"),
    "LSD": ("Ls", "D"), "LSQ": ("Ls", "Q"), "LSRS": ("Ls", "Rs"),"LSRDC": ("Ls", "Rdc"),
    "ZTD": ("Z", "Theta"), "ZTR": ("Z", "Theta"),
    "ZTHD": ("Z", "Theta"), "ZTHR": ("Z", "Theta"),
    "YTD": ("Y", "Theta"), "YTR": ("Y", "Theta"),
    "RX": ("R", "X"),
    "GB": ("G", "B"),
    "VDCIDC": ("Vdc", "Idc"),
}

# ════════════════════════════════════════════════════════════
#  GRAPH AUTO-SCALE  (fixes the live graph collapsing to
#  "0.000000000" for small-magnitude raw SI values)
#
#  ROOT CAUSE of "Cp vs Time shows a blank/flat line at 0": the
#  live graph plots p1_raw / p2_raw straight from the instrument,
#  which are *base SI units* — e.g. a 101.1184 pF capacitance is
#  really the float 1.011184e-10 (Farads). D vs Time looks fine
#  because D is already a human-scale number (~0.138); Cp vs Time
#  does not, because 1e-10 rounded to the formatter's fixed 9
#  decimal places *is* 0.000000000 — the line is really being
#  drawn, just squashed onto the zero line and unreadable.
#
#  Fix: before plotting, auto-pick an engineering divisor/prefix
#  (p, n, µ, m, k, M, ...) from the data's own magnitude — exactly
#  like the front-panel/format_lcr_value already does for the text
#  readout — so the *plotted* numbers land in human-scale range
#  where the adaptive-decimal tick formatter can actually show the
#  variation.
# ════════════════════════════════════════════════════════════
BASE_UNIT = {
    "C": "F", "L": "H",
    "R": "\u03a9", "Z": "\u03a9", "X": "\u03a9",
    "RS": "\u03a9", "RP": "\u03a9",
    "Y": "S", "G": "S", "B": "S",
    "VDC": "V", "IDC": "A",
    "THETA_RAD": "rad",
}

ENG_PREFIXES = [
    (1e-15, "f"), (1e-12, "p"), (1e-9, "n"), (1e-6, "\u00b5"),
    (1e-3, "m"), (1.0, ""), (1e3, "k"), (1e6, "M"), (1e9, "G"),
]


def pick_eng_scale(values):
    """Pick the engineering divisor + prefix that best fits the
    *typical* (median) magnitude of a list of raw values, so a run
    of ~1e-10 F readings scales to ~100 (pF) instead of plotting as
    ~0.0000000001 on a linear axis. Dimensionless/near-unity series
    (D, Q, degrees, radians) naturally fall into the "" (×1) bucket
    and are left untouched."""
    nz = [abs(v) for v in values if v not in (0, None) and
          isinstance(v, (int, float)) and math.isfinite(v)]
    if not nz:
        return 1.0, ""
    nz.sort()
    typical = nz[len(nz) // 2]
    best_div, best_prefix = ENG_PREFIXES[0]
    for div, prefix in ENG_PREFIXES:
        if typical / div >= 1.0:
            best_div, best_prefix = div, prefix
        else:
            break
    return best_div, best_prefix


def get_mode_labels(mode: str):
    """Return (p1_label, p2_label) for a mode name, e.g. 'Cs-Rs' -> ('Cs','Rs')."""
    return MODE_LABELS.get(scpi_code(mode), ("P1", "P2"))
VOLTAGES  = ["0.100", "0.500", "1.000", "2.000"]
VOLT_LABELS = ["0.100 V", "0.500 V", "1.000 V", "2.000 V"]
APERTURES = ["SHORT", "MED", "LONG"]

# Averaging factor selectable for the E4980A (powers of two, matching
# the instrument's actual APER averaging-rate parameter range).
# Sent to the instrument alongside the aperture command, e.g.
# "APER MED,8".
AVERAGES = [str(2 ** i) for i in range(9)]  # "1","2","4",...,"256"
FREQ_PRESETS = [
    ("20 Hz","20"), ("100 Hz","100"), ("1 kHz","1000"),
    ("10 kHz","10000"), ("100 kHz","100000"),
    ("500 kHz","500000"), ("1 MHz","1000000"), ("2 MHz","2000000"),
]

UNIT_ASCII = {
    "µF":"uF","µH":"uH","µS":"uS",
    "Ω":"Ohm","kΩ":"kOhm","MΩ":"MOhm","mΩ":"mOhm","°":"deg",
}
def ascii_safe(t):
    for u, a in UNIT_ASCII.items():
        t = t.replace(u, a)
    return t

def _hz_label(hz):
    try:
        hz = float(hz)
    except (TypeError, ValueError):
        return str(hz)
    if hz >= 1_000_000: return f"{hz/1e6:.4g} MHz"
    if hz >= 1000:      return f"{hz/1e3:.4g} kHz"
    return f"{hz:.4g} Hz"


# ════════════════════════════════════════════════════════════
#  EXPORT SYSTEM  (Phase 1)
#  Standalone, self-contained functions — take a CSV path (and,
#  for the PDF, some run metadata + an optional graph image) and
#  produce a new file alongside it. No dependency on LCRApp state
#  beyond what's explicitly passed in, so these are easy to test
#  and reuse independently of the GUI.
# ════════════════════════════════════════════════════════════
def export_csv_to_excel(csv_path: str) -> str:
    """Read a CSV produced by the measurement loop and write a
    matching .xlsx alongside it (same folder, same base name).
    Bold header row, auto-sized columns. Returns the xlsx path."""
    if not os.path.isfile(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise ValueError("CSV file is empty — nothing to export.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LCR Measurements"

    header, *data_rows = rows
    ws.append(header)
    for row in data_rows:
        # Keep numeric columns numeric in Excel where possible
        # (Reading#, Elapsed(s), Freq(Hz), Volt(V), P1_raw, P2_raw)
        converted = []
        for cell in row:
            try:
                converted.append(float(cell) if "." in cell or
                                 cell.lstrip("-").isdigit() else cell)
            except (ValueError, AttributeError):
                converted.append(cell)
        ws.append(converted)

    # Bold header row
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # Auto-adjust column widths based on the widest cell in each column
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=8)
        ws.column_dimensions[letter].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"  # keep header visible while scrolling

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def export_csv_to_excel_grouped(csv_path: str, title: str = None,
                                duc_title: str = None,
                                temp_rh: str = "") -> str:
    """Read a CSV produced by the measurement loop and write a
    pivoted/grouped .xlsx report: one row per reading, with each
    active mode's two parameters (e.g. Cs/Rs, Cp/D, Z/Theta) laid
    out as adjacent column groups, each split into Value/Unit
    sub-columns where the parameter has a unit (dimensionless ones
    like D or Q get a single column). Matches the lab's preferred
    manually-built report layout: a Date column (from the run's
    Timestamp) split into a Date column and a Time column, plus a
    Temp/RH column — Date and Temp/RH merged down the full
    height of this run's block, exactly like the hand-built master
    sheet — precede the measurement columns. temp_rh is optional
    free text (e.g. "25.86°/52.99%"); leave blank to fill in later
    directly in Excel. Returns the xlsx path (same folder, "_grouped"
    suffix appended so it never collides with the plain export)."""
    if not os.path.isfile(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows or len(rows) < 2:
        raise ValueError("CSV file is empty — nothing to export.")

    header, *data_rows = rows
    idx = {name: i for i, name in enumerate(header)}
    required = ["Reading#", "Mode", "P1_display", "P2_display"]
    if not all(r in idx for r in required):
        # Not the expected long-format measurement CSV — fall back
        # to the simple flat exporter instead of guessing.
        return export_csv_to_excel(csv_path)

    # Date for this whole run, taken from the first row's Timestamp
    # (format DD.MM.YYYY to match the lab's manual report). Falls
    # back to today's date if Timestamp isn't present for some reason.
    run_date = datetime.now().strftime("%d.%m.%Y")
    if "Timestamp" in idx and data_rows:
        raw_ts = data_rows[0][idx["Timestamp"]]
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
            try:
                run_date = datetime.strptime(raw_ts, fmt).strftime("%d.%m.%Y")
                break
            except ValueError:
                continue

    def split_val_unit(disp):
        disp = (disp or "").strip()
        parts = disp.rsplit(" ", 1)
        if len(parts) == 2 and parts[1] and not parts[1].replace(
                ".", "").replace("-", "").isdigit():
            return parts[0], parts[1]
        return disp, None  # None => this parameter has no unit at all

    # Discover the modes present, in first-seen order.
    modes_seen = []
    for row in data_rows:
        m = row[idx["Mode"]]
        if m not in modes_seen:
            modes_seen.append(m)

    # reading# -> {mode: (p1_disp, p2_disp)}
    by_reading = {}
    reading_order = []
    for row in data_rows:
        rno = row[idx["Reading#"]]
        if rno not in by_reading:
            by_reading[rno] = {}
            reading_order.append(rno)
        by_reading[rno][row[idx["Mode"]]] = (
            row[idx["P1_display"]], row[idx["P2_display"]])

    # For each (mode, param position) figure out whether it has a
    # unit at all, by checking the first non-empty sample.
    has_unit = {}  # (mode, 0 or 1) -> bool
    for m in modes_seen:
        for pos in (0, 1):
            sample = next(
                (by_reading[rno][m][pos] for rno in reading_order
                 if m in by_reading[rno] and by_reading[rno][m][pos]),
                "")
            _, u = split_val_unit(sample)
            has_unit[(m, pos)] = u is not None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LCR Report"

    # Blue-themed styling to match the lab's manually-built report
    # (DUC Capacitor sheet): consistent light-blue for both the group
    # header row and the Value/Unit sub-header row, with clean
    # alternating white/gray row banding in place of the old peach tint.
    GROUP_FILL  = PatternFill("solid", fgColor="9DC3E6")
    HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
    ALT_FILL    = PatternFill("solid", fgColor="D9D9D9")
    thin        = Side(style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")

    # Work out column layout: Date (1 col) + Time (1 col) +
    # Temp/RH (1 col) + for
    # each mode, each of its 2 params gets 2 cols (Value, Unit) or
    # 1 col if unitless.
    param_cols = []  # list of (mode, pos, label, width) in column order
    for m in modes_seen:
        p1_label, p2_label = get_mode_labels(m)
        for pos, lbl in ((0, p1_label), (1, p2_label)):
            width = 2 if has_unit[(m, pos)] else 1
            param_cols.append((m, pos, lbl, width))

    total_cols = 3 + sum(w for *_, w in param_cols)

    row_cursor = 1
    if duc_title:
        ws.merge_cells(start_row=row_cursor, start_column=1,
                        end_row=row_cursor, end_column=total_cols)
        dc = ws.cell(row=row_cursor, column=1, value=duc_title)
        dc.font = Font(bold=True, size=14)
        dc.alignment = center
        row_cursor += 1
    if title:
        ws.merge_cells(start_row=row_cursor, start_column=1,
                        end_row=row_cursor, end_column=total_cols)
        c = ws.cell(row=row_cursor, column=1, value=title)
        c.font = Font(bold=True, size=13)
        c.alignment = center
        row_cursor += 2  # blank spacer row after the title

    group_row  = row_cursor
    header_row = row_cursor + 1

    # Date + Temp/RH header cells (each spans both header rows, like
    # the other unitless columns below)
    dcell = ws.cell(row=group_row, column=1, value="Date")
    ws.merge_cells(start_row=group_row, start_column=1,
                    end_row=header_row, end_column=1)
    dcell.font, dcell.alignment, dcell.fill, dcell.border = (
        Font(bold=True), center, GROUP_FILL, border)

    timecell = ws.cell(row=group_row, column=2, value="Time")
    ws.merge_cells(start_row=group_row, start_column=2,
                    end_row=header_row, end_column=2)
    timecell.font, timecell.alignment, timecell.fill, timecell.border = (
        Font(bold=True), center, GROUP_FILL, border)

    trcell = ws.cell(row=group_row, column=3, value="Temp/RH")
    ws.merge_cells(start_row=group_row, start_column=3,
                    end_row=header_row, end_column=3)
    trcell.font, trcell.alignment, trcell.fill, trcell.border = (
        Font(bold=True), center, GROUP_FILL, border)

    col = 4
    col_start = []  # column index each param_cols entry starts at
    for m, pos, lbl, width in param_cols:
        col_start.append(col)
        if width == 2:
            ws.merge_cells(start_row=group_row, start_column=col,
                            end_row=group_row, end_column=col + 1)
            gcell = ws.cell(row=group_row, column=col, value=lbl)
            for cc in range(col, col + 2):
                ws.cell(row=group_row, column=cc).fill = GROUP_FILL
                ws.cell(row=group_row, column=cc).border = border
            for j, sub in enumerate(["Value", "Unit"]):
                hcell = ws.cell(row=header_row, column=col + j, value=sub)
                hcell.font, hcell.alignment = Font(bold=True), center
                hcell.fill, hcell.border = HEADER_FILL, border
        else:
            ws.merge_cells(start_row=group_row, start_column=col,
                            end_row=header_row, end_column=col)
            gcell = ws.cell(row=group_row, column=col, value=lbl)
            gcell.fill, gcell.border = GROUP_FILL, border
        gcell.font, gcell.alignment = Font(bold=True), center
        col += width

    data_start = header_row + 1
    n_rows = len(reading_order)

    # Date + Temp/RH are merged down the entire block of readings in
    # this run, matching the hand-built master sheet where one run
    # (one date/session) shares a single Date + Temp/RH cell.
    if n_rows > 0:
      date_cell = ws.cell(row=data_start, column=1, value=run_date)
      if n_rows > 1:
        ws.merge_cells(start_row=data_start, start_column=1,
                       end_row=data_start+n_rows-1, end_column=1)
      date_cell.alignment, date_cell.border = center, border

      # Temp/RH is one value for the whole run, so it stays merged
      # down the block like Date. Time is per-reading (set in the
      # loop below) so it is deliberately NOT merged.
      tr_cell = ws.cell(row=data_start, column=3, value=temp_rh or "")
      if n_rows > 1:
        ws.merge_cells(start_row=data_start, start_column=3,
                       end_row=data_start+n_rows-1, end_column=3)
      tr_cell.alignment, tr_cell.border = center, border
      

    for i, rno in enumerate(reading_order):
        r = data_start + i

        # Find the row corresponding to this reading
        row_data = next(
          row for row in data_rows
          if row[idx["Reading#"]] == rno
        )

        ts = row_data[idx["Timestamp"]]

        try:
          dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
          dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")

        # Time — unlike Date (merged once, above, across the whole
        # run block), each reading gets its own Time value, so it's
        # written per-row here instead of being merged.
        tcell = ws.cell(row=r, column=2, value=dt.strftime("%H:%M:%S"))
        tcell.alignment = center
        tcell.border = border

        modes_here = by_reading[rno]
        for (m, pos, lbl, width), c0 in zip(param_cols, col_start):
            disp = modes_here.get(m, (None, None))[pos]
            val, unit = split_val_unit(disp) if disp is not None else ("", "")
            try:
                val_out = float(val)
            except (TypeError, ValueError):
                val_out = val
            vcell = ws.cell(row=r, column=c0, value=val_out)
            vcell.alignment, vcell.border = center, border
            if width == 2:
                ucell = ws.cell(row=r, column=c0 + 1, value=unit or "")
                ucell.alignment, ucell.border = center, border

    # Border on the merged Date/Temp-RH cells (only the top-left
    # holder is a real cell; give the whole merged range a border by
    # touching every underlying cell before the merge would also
    # work, but openpyxl only needs the anchor cell styled here since
    # borders on merged ranges render from the anchor + edges below).
    for i in range(n_rows):
        r = data_start + i
        for cc in (1, 3):
            ws.cell(row=r, column=cc).border = border

    # Light banding across this run's whole block (all rows share one
    # tint, since they're one session — alternate exports/appends
    # would naturally land in a different visual block).
    BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
    for i in range(n_rows):
        if i % 2 == 1:
            r = data_start + i
            for cc in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=cc)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BAND_FILL

    for col_idx in range(1, total_cols + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 14  # Temp/RH is wider

    xlsx_path = os.path.splitext(csv_path)[0] + "_grouped.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def export_csv_to_pdf(csv_path: str, meta: dict,
                      graph_image_path: str = None,
                      max_table_rows: int = 200) -> str:
    """Generate a professional PDF report from a measurement CSV.

    meta is expected to contain (all optional, sensible defaults
    used if missing):
        instrument, modes, frequency, voltage, aperture, n_readings

    graph_image_path, if given, is embedded as a chart image.
    Returns the pdf path (same folder, same base name as the CSV).
    """
    if not os.path.isfile(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise ValueError("CSV file is empty — nothing to export.")
    header, *data_rows = rows

    pdf_path = os.path.splitext(csv_path)[0] + ".pdf"
    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        topMargin=18*mm, bottomMargin=16*mm,
        leftMargin=14*mm, rightMargin=14*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=18, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle(
        "ReportSub", parent=styles["Normal"],
        fontSize=10, alignment=TA_CENTER,
        textColor=colors.HexColor("#5B5FFF"), spaceAfter=14)
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=12, spaceBefore=10, spaceAfter=6,
        textColor=colors.HexColor("#1C1C1E"))

    elements = []
    elements.append(Paragraph(
        "LCR Meter Automation — Measurement Report", title_style))
    elements.append(Paragraph(
        "CSIR-NPL LCR Automation Software", sub_style))

    # ── Run metadata table ────────────────────────────────────
    now_str = datetime.now().strftime("%d %B %Y, %I:%M %p")
    meta_rows = [
        ["Report Generated", now_str],
        ["Instrument", meta.get("instrument", "Agilent E4980A")],
        ["Measurement Mode(s)", meta.get("modes", "—")],
        ["Frequency", meta.get("frequency", "—")],
        ["Signal Voltage", meta.get("voltage", "—")],
        ["Aperture", meta.get("aperture", "—")],
        ["Average", meta.get("average", "1")],
        ["Number of Readings", str(meta.get("n_readings",
                                            len(data_rows)))],
        ["Source File", os.path.basename(csv_path)],
    ]
    meta_table = Table(meta_rows, colWidths=[55*mm, 110*mm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#5B5FFF")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E0E0E8")),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 12))

    # ── Embedded graph ─────────────────────────────────────────
    if graph_image_path and os.path.isfile(graph_image_path):
        elements.append(Paragraph("Measurement Graph", section_style))
        # Fit within page width while preserving aspect ratio
        max_w = 170 * mm
        img = RLImage(graph_image_path)
        ratio = img.imageHeight / float(img.imageWidth)
        img.drawWidth = max_w
        img.drawHeight = max_w * ratio
        elements.append(img)
        elements.append(Spacer(1, 10))

    # ── Measurement table ──────────────────────────────────────
    # The full 12-column CSV (including raw scientific-notation
    # floats) is too dense to read on an A4 page at any reasonable
    # font size. The PDF table shows a curated, human-readable
    # subset; the complete raw data remains in the CSV/Excel file.
    elements.append(Paragraph("Measurement Data", section_style))
    truncated = len(data_rows) > max_table_rows

    # Map header names -> column index, then pick a readable subset
    # (falls back to the full set if the expected columns aren't
    # present, so this still works on differently-shaped CSVs).
    col_idx = {name: i for i, name in enumerate(header)}
    preferred = ["Reading#", "Elapsed(s)", "Mode",
                "P1_display", "P2_display", "Status"]
    use_cols = [c for c in preferred if c in col_idx]
    if not use_cols:
        use_cols = header

    display_header = [c.replace("_display", "") for c in use_cols]
    display_rows = [display_header]
    for row in data_rows[:max_table_rows]:
        display_rows.append([row[col_idx[c]] for c in use_cols])

    # Weighted widths: short numeric columns get less room than the
    # mode/status text columns.
    weight_map = {
        "Reading#": 0.7, "Elapsed(s)": 1.0, "Mode": 1.0,
        "P1": 1.6, "P2": 1.6, "Status": 1.1,
    }
    weights = [weight_map.get(c, 1.0) for c in display_header]
    avail_width = 182 * mm
    total_w = sum(weights)
    col_widths = [avail_width * w / total_w for w in weights]

    data_table = Table(display_rows, colWidths=col_widths,
                       repeatRows=1)
    data_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5B5FFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F4FA")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D8D8E4")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(data_table)

    if truncated:
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(
            f"Showing first {max_table_rows} of {len(data_rows)} "
            f"readings. Full data is available in the CSV/Excel file.",
            ParagraphStyle("Note", parent=styles["Normal"],
                          fontSize=8, textColor=colors.HexColor("#8E8E93"))))

    doc.build(elements)
    return pdf_path


# ════════════════════════════════════════════════════════════
#  DUAL LIVE GRAPH WIDGET  (Matplotlib, dark theme)
#  - Stores per-mode time/P1/P2 series (same storage contract
#    as the old LiveGraph, so the measurement thread doesn't
#    need to change: add_point(mode, p1_raw, p2_raw) + clear()).
#  - Renders only ONE selected mode at a time across two
#    stacked graphs: P1 vs Time (cyan) and P2 vs Time (pink).
#  - Titles / Y-axis labels are pulled automatically from
#    formatter.get_mode_labels(mode) -> (p1_name, p2_name).
# ════════════════════════════════════════════════════════════
class DualLiveGraph(tk.Frame):
    def __init__(self, parent, max_points=GRAPH_POINTS, **kwargs):
        super().__init__(parent, bg=BG_GRAPH, **kwargs)
        self.max_points = max_points

        # time_data / p1_data / p2_data, keyed per mode so that
        # switching the dropdown does not lose other modes' history.
        self.time_data: dict[str, deque] = {}
        self.p1_data:   dict[str, deque] = {}
        self.p2_data:   dict[str, deque] = {}

        self.selected_mode = None   # which mode is currently graphed

        # Tracks whether the user has manually zoomed/panned each
        # axis (scroll-wheel or the toolbar's zoom/pan tools). While
        # True for an axis, redraw() stops calling autoscale on it —
        # otherwise every new point would silently snap the view back
        # out and undo the zoom the user just did.
        self._user_zoomed = {"ax1": False, "ax2": False}

        # Per-axis (divisor, engineering-prefix, base-unit) chosen by
        # pick_eng_scale() each redraw — e.g. (1e-12, "p", "F") once
        # Cp readings are running around 100 pF. Used both to label
        # the axis and to recover the true raw value for the
        # full-precision hover readout in _format_coord().
        self._scale = {"ax1": (1.0, "", ""), "ax2": (1.0, "", "")}

        # ── Matplotlib figure: 2 stacked axes ──────────────────
        self.fig = Figure(figsize=(5, 4), dpi=100)
        self.fig.patch.set_facecolor(BG_GRAPH)

        self.ax1 = self.fig.add_subplot(211)
        self.ax2 = self.fig.add_subplot(212)
        self.fig.subplots_adjust(left=0.15, right=0.97,
                                 top=0.90, bottom=0.12,
                                 hspace=0.45)

        for ax in (self.ax1, self.ax2):
            ax.set_facecolor(BG_GRAPH)
            ax.tick_params(colors=T_SECOND, labelsize=8)
            for spine in ax.spines.values():
                spine.set_color(G_GRID)
            ax.grid(True, color=G_GRID, linewidth=0.7,
                   linestyle="-", alpha=0.9)
            # Minor gridlines too — with high-precision formatting the
            # whole point is to be able to zoom in far enough to see
            # every point separated on its own gridline.
            ax.grid(True, which="minor", color=G_GRID,
                   linewidth=0.35, linestyle=":", alpha=0.5)
            ax.minorticks_on()
            ax.set_xlabel("Time (readings)", color=T_SECOND, fontsize=8)
            # Full-precision, non-offset, non-scientific tick labels
            # that scale their decimal count to how far the axis is
            # currently zoomed in — see _make_precise_formatter.
            ax.yaxis.set_major_formatter(
                mticker.FuncFormatter(self._make_precise_formatter(ax)))

        # Full-precision hover readout: replaces the toolbar's default
        # "(x, y) = (...)" bottom-bar text (still %g-limited) with the
        # true raw SI value out to 12 significant figures, so hovering
        # the cursor over any point on the line shows every digit —
        # even variations as small as 0.00000001 that the tick labels
        # themselves round away at a fully zoomed-out view.
        self.ax1.format_coord = lambda x, y: self._format_coord("ax1", x, y)
        self.ax2.format_coord = lambda x, y: self._format_coord("ax2", x, y)

        self.line1, = self.ax1.plot([], [], color=G_CYAN,
                                    linewidth=1.8, marker="o",
                                    markersize=3, label="P1")
        self.line2, = self.ax2.plot([], [], color=G_PINK,
                                    linewidth=1.8, marker="o",
                                    markersize=3, label="P2")

        # ── Always-on min/max/Δ readout ─────────────────────────
        # Direct, unambiguous answer to "is this line really flat,
        # or did zoom just fail to show me the variation?" — this
        # text updates every redraw with the actual spread of the
        # currently plotted data, independent of whatever the Y-axis
        # happens to be zoomed to right now. If Δ reads exactly 0,
        # the readings genuinely haven't changed (not a graph bug);
        # if Δ is nonzero here but the line still looks flat, that's
        # the sign to hit "🔬 Fit to Variation" below.
        spread_row = tk.Frame(self, bg=BG_GRAPH)
        spread_row.pack(side="top", fill="x", padx=6, pady=(2, 0))
        self.lbl_spread1 = tk.Label(spread_row, text="", bg=BG_GRAPH,
                                    fg=G_CYAN, font=F_MONO_SM, anchor="w")
        self.lbl_spread1.pack(side="left", padx=(0, 16))
        self.lbl_spread2 = tk.Label(spread_row, text="", bg=BG_GRAPH,
                                    fg=G_PINK, font=F_MONO_SM, anchor="w")
        self.lbl_spread2.pack(side="left")

        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().configure(bg=BG_GRAPH,
                                              highlightthickness=0)
        self.canvas.get_tk_widget().pack(side="top", fill="both",
                                         expand=True)

        # ── Zoom/pan toolbar + one-click zoom controls ─────────
        tb_row = tk.Frame(self, bg=BG_GRAPH)
        tb_row.pack(side="bottom", fill="x")

        # zoom_group (Auto-Fit, +/-, and — critically — "Fit to
        # Variation") is packed FIRST, before the matplotlib
        # toolbar, so it always gets its space reserved on the
        # right. Previously the toolbar claimed its width first
        # and zoom_group took whatever was left over; on any PC
        # where the row ended up narrower (different Windows
        # display-scaling, smaller screen, etc.) "Fit to Variation"
        # — the last widget in that leftover space — was the one
        # that silently got clipped off-window. Reserving its
        # space first means the *toolbar* icons are what would
        # compress instead, and this button is always visible.
        zoom_group = tk.Frame(tb_row, bg=BG_GRAPH)
        zoom_group.pack(side="right", padx=8, pady=3)

        # The matplotlib toolbar's own zoom (right-click-drag a box,
        # with left-button-pan) is powerful but fiddly to discover.
        # These three buttons are a much easier alternative for
        # "just let me see the small variation": click "＋" repeatedly
        # to tighten the Y-range around the current center — no
        # dragging, no remembering which mouse button does what —
        # "Auto-Fit" to snap back out, "－" to zoom back out a step.
        self.toolbar = NavigationToolbar2Tk(self.canvas, tb_row)
        self.toolbar.update()
        self.toolbar.pack(side="left")

        def _mk_zoom_btn(parent, text, cmd, width=3):
            return tk.Button(parent, text=text, command=cmd,
                             bg="#2A2A40", fg=T_WHITE, relief="flat",
                             font=(SUI, 10, "bold"), width=width,
                             cursor="hand2", activebackground=VIOLET)

        _mk_zoom_btn(zoom_group, "－", lambda: self._zoom(1 / 0.6)
                    ).pack(side="left", padx=(0, 2))
        tk.Button(zoom_group, text="⤢ Auto-Fit", command=self._auto_fit,
                 bg=VIOLET, fg=T_WHITE, relief="flat",
                 font=(SUI, 8, "bold"), padx=10, pady=2,
                 cursor="hand2").pack(side="left", padx=2)
        _mk_zoom_btn(zoom_group, "＋", lambda: self._zoom(0.6)
                    ).pack(side="left", padx=(2, 0))

        # One click straight to "show me the actual variation" —
        # this is the direct answer to "the line just looks flat,
        # can I zoom to see the exact values": instead of clicking
        # "＋" repeatedly and guessing when you've zoomed in enough,
        # this jumps the Y-axis straight to the real min/max spread
        # of the data that's currently plotted, with just a small
        # margin — so a "flat" line snaps immediately into its real
        # zig-zag shape, however tiny the differences are.
        tk.Button(zoom_group, text="🔬 Fit to Variation",
                 command=self._fit_to_variation,
                 bg="#2A2A40", fg=T_WHITE, relief="flat",
                 font=(SUI, 8, "bold"), padx=8, pady=2,
                 cursor="hand2", activebackground=VIOLET
                 ).pack(side="left", padx=(8, 0))

        # Same +/- zoom (and 'v' for Fit to Variation) from the
        # keyboard, once the graph itself has focus — click anywhere
        # on the plot first. Bound to the canvas widget specifically
        # (not bind_all): binding globally would hijack every "+"/"-"
        # keystroke typed anywhere else in the app, e.g. into the
        # Frequency or Voltage entry fields, which is not what we want.
        cw = self.canvas.get_tk_widget()
        cw.bind("<KeyPress-plus>", lambda e: self._zoom(0.6))
        cw.bind("<KeyPress-equal>", lambda e: self._zoom(0.6))
        cw.bind("<KeyPress-minus>", lambda e: self._zoom(1 / 0.6))
        cw.bind("<KeyPress-v>", lambda e: self._fit_to_variation())
        cw.bind("<Button-1>", lambda e: cw.focus_set(), add="+")

        # Scroll-wheel zoom, centered on the cursor, independently on
        # whichever of the two axes the cursor is over.
        self.canvas.mpl_connect("scroll_event", self._on_scroll)
        # Any completed toolbar pan/zoom drag also counts as a
        # manual zoom, so redraw() should stop autoscaling that axis.
        self.canvas.mpl_connect("button_release_event",
                                self._on_button_release)

        self._set_titles(None)

    # ──────────────────────────────────────────────────────────
    #  ZOOM / PAN HELPERS
    # ──────────────────────────────────────────────────────────
    def _make_precise_formatter(self, ax):
        """Y-axis tick formatter whose decimal-place count adapts to
        how far the axis is currently zoomed in, so a tightly zoomed
        view (tiny span) shows enough digits to actually see the
        variation, while a fully zoomed-out view doesn't clutter
        itself with meaningless trailing digits."""
        def fmt(val, pos):
            lo, hi = ax.get_ylim()
            span = abs(hi - lo)
            if span <= 0 or not math.isfinite(span):
                decimals = 6
            else:
                mag = math.floor(math.log10(span))
                # ~5 meaningful digits across the current span
                decimals = max(0, min(12, 4 - mag))
            return f"{val:,.{decimals}f}"
        return fmt

    def _format_coord(self, key, x, y):
        """Full-precision cursor readout for the toolbar's bottom
        status bar. `y` is the plotted (already engineering-scaled)
        value; this converts it back to the true raw SI reading and
        shows both, to 12 significant figures, so no digit the
        instrument reported is ever hidden from view."""
        div, prefix, unit = self._scale.get(key, (1.0, "", ""))
        raw = y * div
        if unit:
            return (f"t = {x:.4g}s    value = {y:,.9f} {prefix}{unit}"
                    f"    (raw = {raw:.12e} {unit})")
        return f"t = {x:.4g}s    value = {y:.12f}"

    def _zoom(self, factor):
        """One-click/keyboard zoom for BOTH graphs at once, centered
        on whatever's currently in view — the easy alternative to the
        matplotlib toolbar's right-click-drag zoom box. factor < 1
        zooms in (tighter Y-range, so tiny variations get taller and
        easier to see); factor > 1 zooms back out. Repeat clicks keep
        tightening the range, so you can zoom in as far as needed to
        see a variation of 0.00000001 pF or any other tiny step —
        the tick formatter and hover readout both keep adding
        decimal places as the range shrinks."""
        for ax, key in ((self.ax1, "ax1"), (self.ax2, "ax2")):
            x0, x1 = ax.get_xlim()
            y0, y1 = ax.get_ylim()
            xc, yc = (x0 + x1) / 2, (y0 + y1) / 2
            ax.set_xlim(xc - (xc - x0) * factor, xc + (x1 - xc) * factor)
            ax.set_ylim(yc - (yc - y0) * factor, yc + (y1 - yc) * factor)
            self._user_zoomed[key] = True
        self.canvas.draw_idle()

    def _spread_text(self, name, vals, prefix, unit):
        """One line of ground truth: the actual min/max/Δ of the
        data currently on that axis, at full precision — independent
        of zoom level. This is what settles "is it really flat or is
        the graph hiding it" without any guessing."""
        unit_str = f"{prefix}{unit}" if unit else ""
        if len(vals) < 2:
            return f"{name}: —"
        lo, hi = min(vals), max(vals)
        delta = hi - lo
        if delta == 0:
            return (f"{name}: {lo:.9f}{(' ' + unit_str) if unit_str else ''}"
                    f"  (Δ = 0 across {len(vals)} readings — "
                    f"genuinely no variation at this precision)")
        return (f"{name}: {lo:.9f}–{hi:.9f}{(' ' + unit_str) if unit_str else ''}"
                f"  (Δ = {delta:.3e}{(' ' + unit_str) if unit_str else ''})")

    def _fit_to_variation(self):
        """Jump the Y-axis straight to the *actual* min/max spread of
        whichever mode is currently plotted, on both graphs — the
        direct one-click answer to "the line looks flat, can I zoom
        in to see the exact values?". Auto-Fit deliberately leaves a
        generous ~5% margin so the whole dataset is comfortably
        inside the frame; that margin is exactly what hides a tiny
        real variation inside a line that then looks perfectly flat.
        This instead pads only a small fraction of the *real* spread,
        so points that differ by 0.00000001 stop looking identical."""
        mode = self.selected_mode
        d1 = list(self.p1_data.get(mode, []))
        d2 = list(self.p2_data.get(mode, []))
        div1, _, _ = self._scale["ax1"]
        div2, _, _ = self._scale["ax2"]

        for ax, raw_vals, div, key in ((self.ax1, d1, div1, "ax1"),
                                       (self.ax2, d2, div2, "ax2")):
            if len(raw_vals) < 2:
                continue
            vals = [v / div for v in raw_vals]
            lo, hi = min(vals), max(vals)
            span = hi - lo
            # Degenerate case: every point is bit-for-bit identical —
            # there's no real spread to fit to, so fall back to a
            # tiny fixed pad around the value instead of a 0-height
            # axis.
            pad = span * 0.15 if span > 0 else max(abs(lo) * 1e-6, 1e-9)
            ax.set_ylim(lo - pad, hi + pad)
            self._user_zoomed[key] = True

        self.canvas.draw_idle()

    def _on_scroll(self, event):
        ax = event.inaxes
        if ax not in (self.ax1, self.ax2):
            return
        if event.xdata is None or event.ydata is None:
            return
        factor = 0.8 if event.button == "up" else 1.25
        x0, x1 = ax.get_xlim()
        y0, y1 = ax.get_ylim()
        xd, yd = event.xdata, event.ydata
        ax.set_xlim(xd - (xd - x0) * factor, xd + (x1 - xd) * factor)
        ax.set_ylim(yd - (yd - y0) * factor, yd + (y1 - yd) * factor)
        key = "ax1" if ax is self.ax1 else "ax2"
        self._user_zoomed[key] = True
        self.canvas.draw_idle()

    def _on_button_release(self, event):
        # Toolbar in "zoom rect" or "pan" mode dragging on either
        # axis — mark that axis as manually zoomed so the next
        # redraw() doesn't snap it back out.
        if self.toolbar.mode and event.inaxes in (self.ax1, self.ax2):
            key = "ax1" if event.inaxes is self.ax1 else "ax2"
            self._user_zoomed[key] = True

    def _auto_fit(self):
        """Reset both axes back to auto-fit-to-data, clearing any
        manual zoom/pan the user had applied."""
        self._user_zoomed["ax1"] = False
        self._user_zoomed["ax2"] = False
        self.ax1.relim(); self.ax1.autoscale_view()
        self.ax2.relim(); self.ax2.autoscale_view()
        self.canvas.draw_idle()

    # ──────────────────────────────────────────────────────────
    #  DATA API  (called from the measurement thread via root.after)
    # ──────────────────────────────────────────────────────────
    def add_point(self, mode, p1_raw, p2_raw, elapsed=None):
        if mode not in self.p1_data:
            self.time_data[mode] = deque(maxlen=self.max_points)
            self.p1_data[mode]   = deque(maxlen=self.max_points)
            self.p2_data[mode]   = deque(maxlen=self.max_points)

        t = elapsed if elapsed is not None else (
            len(self.time_data[mode]))
        self.time_data[mode].append(t)
        self.p1_data[mode].append(p1_raw)
        self.p2_data[mode].append(p2_raw)

    def clear(self):
        self.time_data.clear()
        self.p1_data.clear()
        self.p2_data.clear()
        self.line1.set_data([], [])
        self.line2.set_data([], [])
        self._user_zoomed["ax1"] = False
        self._user_zoomed["ax2"] = False
        self._scale = {"ax1": (1.0, "", ""), "ax2": (1.0, "", "")}
        self.lbl_spread1.config(text="")
        self.lbl_spread2.config(text="")
        self.canvas.draw_idle()

    # ──────────────────────────────────────────────────────────
    #  MODE SELECTION
    # ──────────────────────────────────────────────────────────
    def set_selected_mode(self, mode):
        self.selected_mode = mode
        # Switching modes swaps in a whole different dataset — any
        # zoom that applied to the old parameter's scale no longer
        # makes sense here, so start fresh at auto-fit.
        self._user_zoomed["ax1"] = False
        self._user_zoomed["ax2"] = False
        self._set_titles(mode)

    def _set_titles(self, mode):
        if mode is None:
            p1_name, p2_name = "P1", "P2"
            pt1, pt2 = None, None
        else:
            p1_name, p2_name = get_mode_labels(mode)
            pt1, pt2 = MODE_PARAM_TYPES.get(scpi_code(mode), (None, None))

        # Stashed for redraw(), which needs to know each series'
        # physical parameter type (C, R, D, ...) to decide whether —
        # and by what engineering prefix — to auto-scale it before
        # plotting. See pick_eng_scale()/BASE_UNIT above.
        self._p1_name, self._p2_name = p1_name, p2_name
        self._pt1, self._pt2 = pt1, pt2

        self.ax1.set_title(f"{p1_name} vs Time",
                           color=T_WHITE, fontsize=10,
                           fontweight="bold")
        self.ax1.set_ylabel(p1_name, color=G_CYAN, fontsize=9)

        self.ax2.set_title(f"{p2_name} vs Time",
                           color=T_WHITE, fontsize=10,
                           fontweight="bold")
        self.ax2.set_ylabel(p2_name, color=G_PINK, fontsize=9)

        self.line1.set_label(p1_name)
        self.line2.set_label(p2_name)
        self.ax1.legend(loc="upper right", fontsize=7,
                        facecolor=BG_GRAPH, edgecolor=G_GRID,
                        labelcolor=T_LIGHT)
        self.ax2.legend(loc="upper right", fontsize=7,
                        facecolor=BG_GRAPH, edgecolor=G_GRID,
                        labelcolor=T_LIGHT)

    # ──────────────────────────────────────────────────────────
    #  REDRAW  — plots whichever mode is selected (falls back to
    #  the first active mode if the previous selection isn't active)
    # ──────────────────────────────────────────────────────────
    def redraw(self, active_modes, aperture):
        if not active_modes:
            return

        if self.selected_mode not in active_modes:
            self.set_selected_mode(active_modes[0])

        mode = self.selected_mode
        t  = list(self.time_data.get(mode, []))
        d1 = list(self.p1_data.get(mode, []))
        d2 = list(self.p2_data.get(mode, []))

        if len(t) >= 1:
            # ── Auto-scale into engineering units ──────────────
            # p1_raw/p2_raw are raw SI values (Farads, Ohms, ...).
            # Plotting a run of ~1e-10 F readings directly on a
            # linear axis makes every point round to 0.000000000
            # at any sane number of decimal places — the classic
            # "graph looks empty" symptom. Instead, pick a divisor
            # from the data's own magnitude (pF/nF/mΩ/...) so the
            # plotted numbers land in human-scale range. Dimension-
            # less series (D, Q, angles) aren't in BASE_UNIT and are
            # plotted as-is.
            pt1, pt2 = self._pt1, self._pt2
            div1, prefix1 = (pick_eng_scale(d1) if pt1 in BASE_UNIT
                             else (1.0, ""))
            div2, prefix2 = (pick_eng_scale(d2) if pt2 in BASE_UNIT
                             else (1.0, ""))
            unit1 = BASE_UNIT.get(pt1, "")
            unit2 = BASE_UNIT.get(pt2, "")
            self._scale["ax1"] = (div1, prefix1, unit1)
            self._scale["ax2"] = (div2, prefix2, unit2)

            d1_scaled = [v / div1 for v in d1]
            d2_scaled = [v / div2 for v in d2]

            self.line1.set_data(t, d1_scaled)
            self.line2.set_data(t, d2_scaled)

            self.ax1.set_ylabel(
                f"{self._p1_name} ({prefix1}{unit1})" if unit1
                else self._p1_name, color=G_CYAN, fontsize=9)
            self.ax2.set_ylabel(
                f"{self._p2_name} ({prefix2}{unit2})" if unit2
                else self._p2_name, color=G_PINK, fontsize=9)

            self.lbl_spread1.config(text=self._spread_text(
                self._p1_name, d1_scaled, prefix1, unit1))
            self.lbl_spread2.config(text=self._spread_text(
                self._p2_name, d2_scaled, prefix2, unit2))

            # Only auto-fit an axis the user hasn't manually zoomed —
            # otherwise every new point would yank a zoomed-in view
            # back out to the full range.
            if not self._user_zoomed["ax1"]:
                self.ax1.relim(); self.ax1.autoscale_view()
            if not self._user_zoomed["ax2"]:
                self.ax2.relim(); self.ax2.autoscale_view()

            # Guard against a degenerate (min==max) Y-axis when only
            # one point exists yet — pad it so the line is visible.
            # Use a tiny relative pad (not 10%) so a single point
            # doesn't stretch the axis absurdly wide once more points
            # with real (small) variation start arriving.
            if len(t) == 1:
                for ax, val, zkey in ((self.ax1, d1_scaled[0], "ax1"),
                                      (self.ax2, d2_scaled[0], "ax2")):
                    if self._user_zoomed[zkey]:
                        continue
                    pad = max(abs(val) * 1e-4, 1e-9)
                    ax.set_ylim(val - pad, val + pad)

        self.canvas.draw_idle()



class BodeSweepGraph(tk.Frame):
    """Dedicated Bode-style plot for Frequency Sweep results —
    deliberately separate from DualLiveGraph (which is wired for
    'vs Time' with a single selected mode). This widget plots
    'vs Frequency' on a log X-axis, and shows every active mode
    from the sweep simultaneously rather than one-at-a-time,
    since a sweep is a single bounded run, not a continuous feed."""

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=BG_GRAPH, **kwargs)

        self.fig = Figure(figsize=(5, 4), dpi=100)
        self.fig.patch.set_facecolor(BG_GRAPH)

        self.ax1 = self.fig.add_subplot(211)
        self.ax2 = self.fig.add_subplot(212)
        self.fig.subplots_adjust(left=0.13, right=0.97,
                                 top=0.90, bottom=0.12,
                                 hspace=0.45)

        for ax in (self.ax1, self.ax2):
            ax.set_facecolor(BG_GRAPH)
            ax.set_xscale("log")
            ax.tick_params(colors=T_SECOND, labelsize=8)
            for spine in ax.spines.values():
                spine.set_color(G_GRID)
            ax.grid(True, which="both", color=G_GRID,
                   linewidth=0.6, linestyle="-", alpha=0.8)
            ax.set_xlabel("Frequency (Hz, log scale)",
                         color=T_SECOND, fontsize=8)

        self.ax1.set_title("P1 vs Frequency", color=T_WHITE,
                           fontsize=10, fontweight="bold")
        self.ax2.set_title("P2 vs Frequency", color=T_WHITE,
                           fontsize=10, fontweight="bold")

        self._lines1 = {}   # mode -> Line2D
        self._lines2 = {}

        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().configure(bg=BG_GRAPH,
                                              highlightthickness=0)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def clear(self):
        for ax in (self.ax1, self.ax2):
            ax.clear()
            ax.set_xscale("log")
            ax.set_facecolor(BG_GRAPH)
            ax.tick_params(colors=T_SECOND, labelsize=8)
            for spine in ax.spines.values():
                spine.set_color(G_GRID)
            ax.grid(True, which="both", color=G_GRID,
                   linewidth=0.6, linestyle="-", alpha=0.8)
            ax.set_xlabel("Frequency (Hz, log scale)",
                         color=T_SECOND, fontsize=8)
        self.ax1.set_title("P1 vs Frequency", color=T_WHITE,
                           fontsize=10, fontweight="bold")
        self.ax2.set_title("P2 vs Frequency", color=T_WHITE,
                           fontsize=10, fontweight="bold")
        self._lines1 = {}
        self._lines2 = {}
        self.canvas.draw_idle()

    def set_titles_for_modes(self, modes):
        """Label the Y-axes using the first swept mode's parameter
        names (mixed-mode sweeps still get a usable generic label)."""
        if len(modes) == 1:
            p1_name, p2_name = get_mode_labels(modes[0])
        else:
            p1_name, p2_name = "P1", "P2"
        self.ax1.set_ylabel(p1_name, color=G_CYAN, fontsize=9)
        self.ax2.set_ylabel(p2_name, color=G_PINK, fontsize=9)

    def update_data(self, sweep_data: dict):
        """sweep_data: {mode: {'freqs': [...], 'p1': [...], 'p2': [...]}}
        Redraws every mode's series; called repeatedly as the sweep
        thread reports each new point so the plot fills in live."""
        palette = MODE_COLORS
        for i, (mode, d) in enumerate(sweep_data.items()):
            freqs, p1, p2 = d["freqs"], d["p1"], d["p2"]
            if not freqs:
                continue
            color = palette[i % len(palette)]

            if mode not in self._lines1:
                self._lines1[mode], = self.ax1.plot(
                    [], [], color=color, linewidth=1.6,
                    marker="o", markersize=2.5, label=mode)
                self._lines2[mode], = self.ax2.plot(
                    [], [], color=color, linewidth=1.6,
                    marker="o", markersize=2.5, label=mode)

            self._lines1[mode].set_data(freqs, p1)
            self._lines2[mode].set_data(freqs, p2)

        if self._lines1:
            self.ax1.relim(); self.ax1.autoscale_view()
            self.ax2.relim(); self.ax2.autoscale_view()
            self.ax1.legend(loc="upper right", fontsize=7,
                            facecolor=BG_GRAPH, edgecolor=G_GRID,
                            labelcolor=T_LIGHT)
            self.ax2.legend(loc="upper right", fontsize=7,
                            facecolor=BG_GRAPH, edgecolor=G_GRID,
                            labelcolor=T_LIGHT)
        self.canvas.draw_idle()


# ════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ════════════════════════════════════════════════════════════
class LCRApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("CSIR-NPL LCR Automation Software")
        self.root.configure(bg=BG_APP)
        self.root.geometry("1350x820")
        self.root.minsize(1100, 720)
        self.root.resizable(True, True)

        # Tkinter normally prints exceptions raised inside .after()
        # callbacks (which is how every live reading reaches the
        # screen — see _measure_loop's root.after(0, self._update_display,
        # ...)) to stderr and otherwise carries on silently. In a
        # windowed/--noconsole EXE there is no stderr for the user to
        # see, so a single bad callback freezes the readings panel
        # with zero visible error — exactly "looks idle, no errors
        # shown". Route those exceptions to the on-screen log AND a
        # file next to the app so they're never invisible again.
        self.root.report_callback_exception = self._handle_tk_callback_exception

        # ── State (100% preserved) ────────────────────────────
        self.lcr            = None
        self.running        = False
        self.paused         = False
        # Tracks which SCPI mode code the instrument's free-run
        # buffer was last force-settled for (see _synchronized_fetch).
        # Reset to None at the start of every run so the first
        # reading of a new run never trusts a leftover buffer value
        # from a previous run/mode.
        self._current_synced_mode = None
        self.pause_event    = threading.Event()
        self.pause_event.set()   # set = run freely; cleared = paused
        self._pause_started_at = None
        self._paused_seconds   = 0.0
        self.reading_count  = 0
        self.start_time     = None
        self.csv_writer     = None
        self.csv_file       = None
        self.csv_path       = tk.StringVar(value="")
        self._stats_data: dict[str, list] = {}
        self._idn_str       = ""

        # ── Data Logger page state (Statistics / Data Logger tab) ──
        # Rolling window of the most recent readings (any mode),
        # newest last — used to populate the live readings table.
        self.readings_log = deque(maxlen=DATALOGGER_MAX_ROWS)
        # One summary entry per completed run, appended in _finish().
        self.session_history = []
        self._selected_session_idx = None
        self._dl_hist_mode = tk.StringVar(value="")
        self._autosave_path = os.path.join(
            os.getcwd(), AUTOSAVE_FILE)

        # ── Open/Short calibration state ──────────────────────
        # Carries over across reconnects within the same app session
        # (physically disconnecting GPIB doesn't change the test
        # fixture, so re-cal isn't forced on reconnect) — only reset
        # when the user explicitly clicks "Clear Calibration".
        self.open_cal_done  = False
        self.short_cal_done = False
        self.calibration_file = "calibration_status.json"
        
        self.calibrating    = False

        # ── Export state ──────────────────────────────────────
        # Export is only enabled once a measurement run has fully
        # completed — exporting a CSV that's still being written,
        # or one that never existed, would be misleading.
        self.last_completed_run = None  # dict of run metadata, or None

        # ── Frequency Sweep state ──────────────────────────────
        self.sweeping        = False
        self.sweep_csv_writer = None
        self.sweep_csv_file   = None
        self.sweep_data: dict[str, dict] = {}   # mode -> {freqs, p1, p2}

        # ── DC Bias state ──────────────────────────────────────
        self.bias_busy = False   # true only while a SCPI bias call is in flight
        self.bias_on   = False

        # ── Control vars (100% preserved) ────────────────────
        self.var_freq_hz   = tk.StringVar(value="100000")
        self.var_volt_var  = tk.StringVar(value="1.000")
        self.var_aperture  = tk.StringVar(value="MED")
        self.var_average   = tk.StringVar(value="1")
        self.var_n         = tk.StringVar(value="100")
        self.var_interval  = tk.StringVar(value="1.0")

        # Frequency Sweep inputs
        self.var_sweep_start = tk.StringVar(value="1000")
        self.var_sweep_stop  = tk.StringVar(value="1000000")
        self.var_sweep_steps = tk.StringVar(value="20")

        # DC Bias inputs
        self.var_bias_volt    = tk.StringVar(value="0.0")
        self.var_bias_enabled = tk.BooleanVar(value=False)
        self.selected_modes = [
            tk.BooleanVar(value=(m == "CPD")) for m in MODES
        ]

        # ── Build UI ──────────────────────────────────────────
        self._build_ui()
        self.load_calibration()
        self._tick_clock()
        self._refresh_cal_status()   # Start disabled until Open+Short cal done
        self.log("Ready — click 'Connect Instrument' to begin.")
        self._autosave_tick()
        self.root.after(400, self._offer_session_reload)

    # ──────────────────────────────────────────────────────────
    #  CLOCK TICKER
    # ──────────────────────────────────────────────────────────
    def _tick_clock(self):
        now = datetime.now()
        self.lbl_date.config(text=now.strftime("%d %B %Y"))
        self.lbl_time.config(text=now.strftime("%I:%M:%S %p"))
        self.root.after(1000, self._tick_clock)

    # ──────────────────────────────────────────────────────────
    #  TOP-LEVEL LAYOUT
    # ──────────────────────────────────────────────────────────
    def _build_ui(self):
        # Root grid: col0=sidebar, col1=main
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main()

    # ──────────────────────────────────────────────────────────
    #  SIDEBAR
    # ──────────────────────────────────────────────────────────
    def _build_sidebar(self):
        sb = tk.Frame(self.root, bg=BG_SIDEBAR, width=150)
        sb.grid(row=0, column=0, sticky="nsew")
        sb.grid_propagate(False)

        # Logo area
        logo = tk.Frame(sb, bg=BG_SIDEBAR, pady=20)
        logo.pack(fill="x")
        # Circular icon placeholder
        icon_c = tk.Canvas(logo, width=56, height=56,
                           bg=BG_SIDEBAR, highlightthickness=0)
        icon_c.pack()
        icon_c.create_oval(4, 4, 52, 52, fill=GRAD_MID, outline="")
        icon_c.create_text(28, 28, text="⚡", fill=T_WHITE,
                           font=(SUI, 22, "bold"))

        tk.Frame(sb, bg="#2A2A3E", height=1).pack(fill="x", pady=6)

        # Nav items
        nav_items = [
            ("🏠", "Dashboard",   True,  self._nav_go_dashboard),
            ("📊", "Measurement", False, self._nav_go_dashboard),
            ("📋", "Data Logger", False, self._nav_go_datalogger),
            ("📈", "Graph",       False, self._nav_go_graph),
            ("💾", "Export",      False, None),
        ]
        self._nav_btns = []
        self._nav_widgets = []   # (frame, icon_label, text_label)
        for icon, label, active, handler in nav_items:
            f = tk.Frame(sb, bg=GRAD_MID if active else BG_SIDEBAR,
                         cursor="hand2")
            f.pack(fill="x", pady=1)
            lbl_icon = tk.Label(f, text=icon,
                     bg=GRAD_MID if active else BG_SIDEBAR,
                     fg=T_WHITE, font=(SUI, 14),
                     pady=8)
            lbl_icon.pack()
            lbl_text = tk.Label(f, text=label,
                     bg=GRAD_MID if active else BG_SIDEBAR,
                     fg=T_WHITE if active else T_SECOND,
                     font=F_NAV, pady=2)
            lbl_text.pack()
            self._nav_btns.append(f)
            self._nav_widgets.append((f, lbl_icon, lbl_text))

            if handler is not None:
                idx = len(self._nav_widgets) - 1
                for w in (f, lbl_icon, lbl_text):
                    w.bind("<Button-1>",
                          lambda e, h=handler, i=idx: self._on_nav_click(h, i))

        # Bottom connect btn
        tk.Frame(sb, bg=BG_SIDEBAR).pack(fill="both", expand=True)
        conn_f = tk.Frame(sb, bg=BG_SIDEBAR, pady=8)
        conn_f.pack(fill="x")
        self.btn_connect = tk.Button(
            conn_f,
            text="Connect\nInstrument",
            command=self._connect,
            bg=VIOLET, fg=T_WHITE,
            font=F_NAV, relief="flat",
            activebackground=GRAD_END,
            cursor="hand2", pady=8, bd=0,
            wraplength=120)
        self.btn_connect.pack(fill="x", padx=10)

    # ──────────────────────────────────────────────────────────
    #  MAIN AREA
    # ──────────────────────────────────────────────────────────
    def _build_main(self):
        main = tk.Frame(self.root, bg=BG_APP)
        main.grid(row=0, column=1, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        self._build_header(main)
        self._build_body(main)

    # ──────────────────────────────────────────────────────────
    #  HEADER
    # ──────────────────────────────────────────────────────────
    def _build_header(self, parent):
        hdr = tk.Frame(parent, bg=BG_HEADER, height=80)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        hdr.columnconfigure(1, weight=1)

        # Title block
        t_blk = tk.Frame(hdr, bg=BG_HEADER)
        t_blk.grid(row=0, column=0, padx=20, pady=12, sticky="w")
        tk.Label(t_blk, text=APP_TITLE,
                 bg=BG_HEADER, fg=T_WHITE,
                 font=F_TITLE).pack(anchor="w")
        tk.Label(t_blk, text=APP_SUB,
                 bg=BG_HEADER, fg=T_SECOND,
                 font=F_SUBTITLE).pack(anchor="w")

        # Connection pill (centre)
        conn_f = tk.Frame(hdr, bg="#1E1E30",
                          padx=14, pady=8)
        conn_f.grid(row=0, column=1, padx=20, sticky="")
        conn_inner = tk.Frame(conn_f, bg="#1E1E30")
        conn_inner.pack()
        self.lbl_conn_dot = tk.Label(
            conn_inner, text="●",
            bg="#1E1E30", fg=T_SECOND,
            font=(SUI, 12))
        self.lbl_conn_dot.pack(side="left")
        self.lbl_conn_text = tk.Label(
            conn_inner, text="  Not Connected",
            bg="#1E1E30", fg=T_SECOND,
            font=F_LABEL_B)
        self.lbl_conn_text.pack(side="left")
        tk.Label(conn_f, text=GPIB_ADDRESS,
                 bg="#1E1E30", fg=T_DIM,
                 font=F_TINY).pack()

        # Date/time block (right)
        dt_f = tk.Frame(hdr, bg="#1E1E30", padx=14, pady=8)
        dt_f.grid(row=0, column=2, padx=20, sticky="e")
        dt_inner = tk.Frame(dt_f, bg="#1E1E30")
        dt_inner.pack()
        tk.Label(dt_inner, text="📅 ",
                 bg="#1E1E30", fg=T_SECOND,
                 font=(SUI, 13)).pack(side="left")
        dt_text = tk.Frame(dt_inner, bg="#1E1E30")
        dt_text.pack(side="left")
        self.lbl_date = tk.Label(dt_text, text="",
                                  bg="#1E1E30", fg=T_WHITE,
                                  font=F_BODY_B)
        self.lbl_date.pack(anchor="w")
        self.lbl_time = tk.Label(dt_text, text="",
                                  bg="#1E1E30", fg=T_SECOND,
                                  font=F_SMALL)
        self.lbl_time.pack(anchor="w")

        tk.Frame(parent, bg="#2A2A3E", height=1).grid(
            row=0, column=0, sticky="ew", pady=(79, 0))

    # ──────────────────────────────────────────────────────────
    #  BODY
    # ──────────────────────────────────────────────────────────
    def _build_body(self, parent):
        # ── Scrollable container ───────────────────────────────
        # The dashboard (modes + settings + live readings + graph)
        # can be taller than the window, especially on smaller
        # screens. Wrap everything in a Canvas+Scrollbar so nothing
        # is silently clipped, and so sidebar nav clicks (e.g.
        # "Graph") can scroll a specific section into view.
        scroll_outer = tk.Frame(parent, bg=BG_APP)
        scroll_outer.grid(row=1, column=0, sticky="nsew")
        scroll_outer.rowconfigure(0, weight=1)
        scroll_outer.columnconfigure(0, weight=1)

        self.body_canvas = tk.Canvas(
            scroll_outer, bg=BG_APP, highlightthickness=0, bd=0)
        self.body_canvas.grid(row=0, column=0, sticky="nsew")

        vsb = ttk.Scrollbar(scroll_outer, orient="vertical",
                            command=self.body_canvas.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.body_canvas.configure(yscrollcommand=vsb.set)

        body = tk.Frame(self.body_canvas, bg=BG_APP, padx=14, pady=12)
        self._body_window = self.body_canvas.create_window(
            (0, 0), window=body, anchor="nw")

        def _on_body_configure(event):
            self.body_canvas.configure(
                scrollregion=self.body_canvas.bbox("all"))
        body.bind("<Configure>", _on_body_configure)

        def _on_canvas_configure(event):
            # Keep the inner frame exactly as wide as the canvas
            # so the two-column layout still resizes correctly.
            self.body_canvas.itemconfig(
                self._body_window, width=event.width)
        self.body_canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            self.body_canvas.yview_scroll(int(-event.delta / 120), "units")

        def _bind_wheel(event):
            self.body_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_wheel(event):
            self.body_canvas.unbind_all("<MouseWheel>")

        # Only capture the wheel while the cursor is actually over
        # the dashboard area — avoids hijacking scroll on widgets
        # added by future features elsewhere in the app.
        self.body_canvas.bind("<Enter>", _bind_wheel)
        self.body_canvas.bind("<Leave>", _unbind_wheel)

        body.columnconfigure(0, weight=0, minsize=430)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # ── Left column ───────────────────────────────────────
        left = tk.Frame(body, bg=BG_APP)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(1, weight=1)

        self._build_modes_card(left)
        self._build_settings_card(left)
        self._build_sweep_card(left)
        self._build_bias_card(left)
        self._build_instrument_card(left)

        # ── Right column ──────────────────────────────────────
        right = tk.Frame(body, bg=BG_APP)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=2)
        right.rowconfigure(2, weight=1)
        right.columnconfigure(0, weight=1)

        self._build_live_readings(right)
        self._build_graph_area(right)
        self._build_sweep_graph_area(right)

        # ── Data Logger / Statistics row (full width, below both
        #    columns) ─────────────────────────────────────────
        body.rowconfigure(1, weight=0)
        self._build_datalogger_card(body)

    # ──────────────────────────────────────────────────────────
    #  SIDEBAR NAV — SCROLL TARGETS
    # ──────────────────────────────────────────────────────────
    def _on_nav_click(self, handler, active_idx):
        for i, (f, lbl_icon, lbl_text) in enumerate(self._nav_widgets):
            is_active = (i == active_idx)
            bg = GRAD_MID if is_active else BG_SIDEBAR
            f.config(bg=bg)
            lbl_icon.config(bg=bg)
            lbl_text.config(bg=bg, fg=T_WHITE if is_active else T_SECOND)
        handler()

    def _nav_go_dashboard(self):
        """Scroll the body canvas back to the very top."""
        self.body_canvas.yview_moveto(0.0)

    def _nav_go_graph(self):
        """Scroll the body canvas so the Live Graph card is in
        view, then flash its border so it's obvious something
        happened (the graphs themselves are unchanged — this
        just brings them on-screen)."""
        self.root.update_idletasks()
        bbox = self.body_canvas.bbox("all")
        if not bbox or not hasattr(self, "graph_card_outer"):
            return
        _, _, _, total_h = bbox

        # winfo_y() is only relative to the immediate parent, which
        # breaks if the layout nesting ever changes. Root-coordinate
        # difference is robust regardless of how deeply the card is
        # nested inside the scrollable body.
        target_y = (self.graph_card_outer.winfo_rooty()
                   - self.body_canvas.winfo_rooty()
                   + self.body_canvas.canvasy(0))

        if total_h > 0:
            # Leave a little headroom above the card instead of
            # pinning it to the very top edge of the viewport.
            frac = max(0.0, (target_y - 16) / total_h)
            self.body_canvas.yview_moveto(frac)
        self._flash_graph_card()

    def _flash_graph_card(self, _step=0):
        """Brief highlight pulse on the graph card border to draw
        the eye after a scroll-to-section jump."""
        flash_colors = [CYAN, "#D0D0E0"]
        if _step >= len(flash_colors) * 2:
            self.graph_card_outer.config(bg="#D0D0E0")
            return
        color = flash_colors[_step % 2]
        self.graph_card_outer.config(bg=color)
        self.root.after(160, lambda: self._flash_graph_card(_step + 1))

    def _nav_go_datalogger(self):
        """Scroll the body canvas so the Data Logger card is in
        view, then flash its border (same pattern as Graph nav)."""
        self.root.update_idletasks()
        bbox = self.body_canvas.bbox("all")
        if not bbox or not hasattr(self, "datalogger_card_outer"):
            return
        _, _, _, total_h = bbox
        target_y = (self.datalogger_card_outer.winfo_rooty()
                   - self.body_canvas.winfo_rooty()
                   + self.body_canvas.canvasy(0))
        if total_h > 0:
            frac = max(0.0, (target_y - 16) / total_h)
            self.body_canvas.yview_moveto(frac)
        self._flash_datalogger_card()

    def _flash_datalogger_card(self, _step=0):
        flash_colors = [CYAN, "#D0D0E0"]
        if _step >= len(flash_colors) * 2:
            self.datalogger_card_outer.config(bg="#D0D0E0")
            return
        color = flash_colors[_step % 2]
        self.datalogger_card_outer.config(bg=color)
        self.root.after(
            160, lambda: self._flash_datalogger_card(_step + 1))

    # ──────────────────────────────────────────────────────────
    #  CARD HELPER
    # ──────────────────────────────────────────────────────────
    def _white_card(self, parent, title, icon="", **grid_kwargs):
        """White card with violet section title, subtle shadow effect."""
        outer = tk.Frame(parent, bg="#D0D0E0",
                         padx=1, pady=1)
        outer.pack(**grid_kwargs)
        inner = tk.Frame(outer, bg=BG_WHITE)
        inner.pack(fill="both", expand=True)

        if title:
            hdr = tk.Frame(inner, bg=BG_WHITE, pady=12, padx=16)
            hdr.pack(fill="x")
            tk.Label(hdr, text=f"{icon}  {title}" if icon else title,
                     bg=BG_WHITE, fg=T_DARK,
                     font=F_SECTION).pack(side="left")
            tk.Frame(inner, bg="#EBEBF0", height=1).pack(fill="x")

        return inner

    # ──────────────────────────────────────────────────────────
    #  MEASUREMENT MODES CARD
    # ──────────────────────────────────────────────────────────
    def _build_modes_card(self, parent):
        card = self._white_card(parent, "Measurement Modes", "📊",
                                fill="x", pady=(0, 8))

        self.lbl_mode_warn = tk.Label(
            card, text="",
            bg=BG_WHITE, fg=RED,
            font=F_SMALL)
        self.lbl_mode_warn.pack(fill="x", padx=16)

        grid = tk.Frame(card, bg=BG_WHITE, padx=12, pady=8)
        grid.pack(fill="x")

        COLS = 4
        self._mode_btns = []
        for i, mode in enumerate(MODES):
            r, c_col = divmod(i, COLS)
            bg  = MODE_CARD_COLORS[i % len(MODE_CARD_COLORS)]
            fg  = MODE_CARD_FG[i % len(MODE_CARD_FG)]
            btn = tk.Button(
                grid, text=mode,
                command=lambda idx=i: self._toggle_mode(idx),
                bg=bg, fg=fg,
                font=(SUI, 9, "bold"),
                relief="flat", bd=0,
                padx=6, pady=6,
                cursor="hand2",
                activebackground=bg)
            btn.grid(row=r, column=c_col,
                     padx=3, pady=3, sticky="ew")
            self._mode_btns.append(btn)
            grid.columnconfigure(c_col, weight=1)
        self._refresh_mode_styles()

        tk.Frame(card, bg=BG_WHITE, height=6).pack()

    # ──────────────────────────────────────────────────────────
    #  SETTINGS CARD
    # ──────────────────────────────────────────────────────────
    def _build_settings_card(self, parent):
        card = self._white_card(parent, "Measurement Settings", "⚙️",
                                fill="x", pady=(0, 8))

        def setting_row(icon, label, widget_fn):
            row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=5)
            row.pack(fill="x")
            tk.Label(row, text=icon, bg=BG_WHITE,
                     fg=VIOLET, font=(SUI, 12),
                     width=2).pack(side="left")
            tk.Label(row, text=label,
                     bg=BG_WHITE, fg=T_DARK,
                     font=F_BODY, width=16,
                     anchor="w").pack(side="left", padx=(6, 0))
            widget_fn(row)
            tk.Frame(card, bg="#F0F0F5", height=1).pack(
                fill="x", padx=16)

        # Frequency row
        def freq_widget(row):
            f = tk.Frame(row, bg=BG_WHITE)
            f.pack(side="right", fill="x", expand=True)
            ent_f = tk.Frame(f, bg="#EEF0FF", padx=8, pady=4)
            ent_f.pack(side="left")
            self.ent_freq = tk.Entry(
                ent_f,
                textvariable=self.var_freq_hz,
                bg="#EEF0FF", fg=T_DARK,
                insertbackground=T_DARK,
                font=(SUI, 11, "bold"),
                relief="flat", width=10, bd=0)
            self.ent_freq.pack(side="left")
            tk.Label(ent_f, text="Hz",
                     bg="#EEF0FF", fg=T_SECOND,
                     font=F_LABEL).pack(side="left", padx=(4, 0))
            self.lbl_freq_hint = tk.Label(
                f, text="= 100 kHz",
                bg=BG_WHITE, fg=GREEN,
                font=F_SMALL)
            self.lbl_freq_hint.pack(side="left", padx=8)
            self.var_freq_hz.trace_add("write", self._validate_freq)
            # Quick chips below
            chips = tk.Frame(card, bg=BG_WHITE, padx=16, pady=4)
            chips.pack(fill="x")
            for lbl, val in FREQ_PRESETS:
                tk.Button(
                    chips, text=lbl,
                    command=lambda v=val: self.var_freq_hz.set(v),
                    bg="#EEF0FF", fg=VIOLET,
                    font=(SUI, 8), relief="flat",
                    cursor="hand2", padx=5, pady=2, bd=0
                ).pack(side="left", padx=2)
            tk.Frame(card, bg="#F0F0F5", height=1).pack(
                fill="x", padx=16)

        setting_row("〜", "Frequency", freq_widget)

        # Signal Voltage — dropdown
        def volt_widget(row):
            cb = ttk.Combobox(
                row,
                textvariable=self.var_volt_var,
                values=VOLT_LABELS,
                state="readonly", width=12,
                font=F_BODY)
            cb.pack(side="right")
            # Map display value back to raw
            def _volt_selected(e):
                val = cb.get().replace(" V", "").strip()
                self.var_volt_var.set(val)
            cb.bind("<<ComboboxSelected>>", _volt_selected)
            cb.set("1.000 V")
        setting_row("⚡", "Signal Voltage", volt_widget)

        # Aperture
        def aper_widget(row):
            cb = ttk.Combobox(
                row,
                textvariable=self.var_aperture,
                values=APERTURES,
                state="readonly", width=12,
                font=F_BODY)
            cb.pack(side="right")
        setting_row("🕐", "Aperture Time", aper_widget)

        # Average (averaging factor sent to the instrument alongside
        # the aperture command — 0 disables extra averaging)
        def avg_widget(row):
            cb = ttk.Combobox(
                row,
                textvariable=self.var_average,
                values=AVERAGES,
                state="readonly", width=12,
                font=F_BODY)
            cb.pack(side="right")
        setting_row("📊", "Average", avg_widget)

        # No. of readings
        def n_widget(row):
            f = tk.Frame(row, bg="#EEF0FF", padx=8, pady=4)
            f.pack(side="right")
            tk.Entry(
                f,
                textvariable=self.var_n,
                bg="#EEF0FF", fg=T_DARK,
                insertbackground=T_DARK,
                font=(SUI, 11, "bold"),
                relief="flat", width=8, bd=0
            ).pack()
        setting_row("🔢", "No. of Readings", n_widget)

        # Interval
        def int_widget(row):
            f = tk.Frame(row, bg="#EEF0FF", padx=8, pady=4)
            f.pack(side="right")
            inner = tk.Frame(f, bg="#EEF0FF")
            inner.pack()
            tk.Entry(
                inner,
                textvariable=self.var_interval,
                bg="#EEF0FF", fg=T_DARK,
                insertbackground=T_DARK,
                font=(SUI, 11, "bold"),
                relief="flat", width=6, bd=0
            ).pack(side="left")
            tk.Label(inner, text=" sec",
                     bg="#EEF0FF", fg=T_SECOND,
                     font=F_LABEL).pack(side="left")
        setting_row("⏱️", "Sampling Interval", int_widget)

        # ── Open / Short Calibration row ──────────────────────
        cal_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=8)
        cal_row.pack(fill="x")
        tk.Label(cal_row, text="🧪", bg=BG_WHITE,
                fg=VIOLET, font=(SUI, 12), width=2).pack(side="left")
        tk.Label(cal_row, text="Calibration",
                bg=BG_WHITE, fg=T_DARK,
                font=F_BODY, width=16,
                anchor="w").pack(side="left", padx=(6, 0))

        cal_btns = tk.Frame(cal_row, bg=BG_WHITE)
        cal_btns.pack(side="right")

        self.btn_open_cal = tk.Button(
            cal_btns, text="Open Cal  ○",
            command=self._run_open_cal,
            bg="#EEF0FF", fg=VIOLET,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=8, pady=4, bd=0)
        self.btn_open_cal.pack(side="left", padx=(0, 4))

        self.btn_short_cal = tk.Button(
            cal_btns, text="Short Cal  ○",
            command=self._run_short_cal,
            bg="#EEF0FF", fg=VIOLET,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=8, pady=4, bd=0)
        self.btn_short_cal.pack(side="left", padx=(0, 4))

        self.btn_clear_cal = tk.Button(
            cal_btns, text="Clear",
            command=self._clear_calibration,
            bg="#FFF0F0", fg=RED,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=8, pady=4, bd=0)
        self.btn_clear_cal.pack(side="left")

        tk.Frame(card, bg="#F0F0F5", height=1).pack(
            fill="x", padx=16)

        # Calibration status / warning line
        cal_status_row = tk.Frame(card, bg=BG_WHITE, padx=16)
        cal_status_row.pack(fill="x", pady=(2, 6))
        self.lbl_cal_status = tk.Label(
            cal_status_row,
            text="⚠  Calibration required before Start is enabled",
            bg=BG_WHITE, fg=AMBER,
            font=F_SMALL, anchor="w", justify="left",
            wraplength=380)
        self.lbl_cal_status.pack(side="left", fill="x", expand=True)

        # CSV row
        csv_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=8)
        csv_row.pack(fill="x")
        tk.Label(csv_row, text="💾  Output File",
                 bg=BG_WHITE, fg=T_DARK,
                 font=F_BODY).pack(side="left")
        tk.Button(
            csv_row, text="Browse",
            command=self._browse,
            bg="#EEF0FF", fg=VIOLET,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=8, pady=3, bd=0
        ).pack(side="right")
        tk.Entry(
            csv_row,
            textvariable=self.csv_path,
            bg="#F8F8FF", fg=T_SECOND,
            font=F_SMALL, relief="flat", bd=2
        ).pack(side="right", fill="x", expand=True,
               padx=(8, 6))

        tk.Frame(card, bg="#F0F0F5", height=1).pack(
            fill="x", padx=16)

        # ── Export row ──────────────────────────────────────
        export_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=8)
        export_row.pack(fill="x")
        tk.Label(export_row, text="📤  Export Results",
                bg=BG_WHITE, fg=T_DARK,
                font=F_BODY).pack(side="left")
        self.btn_export = tk.Button(
            export_row, text="Export",
            command=self._export_clicked,
            bg=VIOLET, fg=T_WHITE,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=14, pady=4, bd=0)
        self.btn_export.pack(side="right")
        self.lbl_export_hint = tk.Label(
            export_row, text="(run a measurement first)",
            bg=BG_WHITE, fg=T_SECOND, font=F_TINY)
        self.lbl_export_hint.pack(side="right", padx=(0, 8))

        tk.Frame(card, bg=BG_WHITE, height=4).pack()

    # ──────────────────────────────────────────────────────────
    #  FREQUENCY SWEEP CARD  (Phase 2)
    # ──────────────────────────────────────────────────────────
    def _build_sweep_card(self, parent):
        card = self._white_card(parent, "Frequency Sweep", "📡",
                                fill="x", pady=(0, 8))

        def sweep_row(icon, label, widget_fn):
            row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=5)
            row.pack(fill="x")
            tk.Label(row, text=icon, bg=BG_WHITE,
                     fg=VIOLET, font=(SUI, 12),
                     width=2).pack(side="left")
            tk.Label(row, text=label,
                     bg=BG_WHITE, fg=T_DARK,
                     font=F_BODY, width=16,
                     anchor="w").pack(side="left", padx=(6, 0))
            widget_fn(row)
            tk.Frame(card, bg="#F0F0F5", height=1).pack(
                fill="x", padx=16)

        def entry_widget(var, width=10, suffix=""):
            def _build(row):
                f = tk.Frame(row, bg="#EEF0FF", padx=8, pady=4)
                f.pack(side="right")
                tk.Entry(
                    f, textvariable=var,
                    bg="#EEF0FF", fg=T_DARK,
                    insertbackground=T_DARK,
                    font=(SUI, 11, "bold"),
                    relief="flat", width=width, bd=0
                ).pack(side="left")
                if suffix:
                    tk.Label(f, text=suffix, bg="#EEF0FF",
                            fg=T_SECOND, font=F_LABEL
                    ).pack(side="left", padx=(4, 0))
            return _build

        sweep_row("〜", "Start Frequency",
                  entry_widget(self.var_sweep_start, suffix="Hz"))
        sweep_row("〜", "Stop Frequency",
                  entry_widget(self.var_sweep_stop, suffix="Hz"))
        sweep_row("🔢", "No. of Steps",
                  entry_widget(self.var_sweep_steps, width=8))

        # Buttons
        btn_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=10)
        btn_row.pack(fill="x")
        self.btn_sweep_start = tk.Button(
            btn_row, text="▶  Start Sweep",
            command=self._start_sweep,
            bg=VIOLET, fg=T_WHITE,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=12, pady=6, bd=0)
        self.btn_sweep_start.pack(side="left", padx=(0, 6))
        self.btn_sweep_stop = tk.Button(
            btn_row, text="■  Stop Sweep",
            command=self._stop_sweep,
            bg=RED, fg=T_WHITE,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=12, pady=6, bd=0,
            state="disabled")
        self.btn_sweep_stop.pack(side="left")

        # Progress
        prog_row = tk.Frame(card, bg=BG_WHITE, padx=16)
        prog_row.pack(fill="x", pady=(0, 10))
        style = ttk.Style()
        style.configure("Sweep.Horizontal.TProgressbar",
                        troughcolor="#EEF0FF", bordercolor="#EEF0FF",
                        background=CYAN, lightcolor=CYAN,
                        darkcolor=CYAN, thickness=8)
        self.sweep_progressbar = ttk.Progressbar(
            prog_row, style="Sweep.Horizontal.TProgressbar",
            orient="horizontal", mode="determinate",
            maximum=100, value=0)
        self.sweep_progressbar.pack(fill="x")
        self.lbl_sweep_status = tk.Label(
            card, text="Idle — set start/stop frequency and steps",
            bg=BG_WHITE, fg=T_SECOND, font=F_SMALL,
            padx=16, anchor="w")
        self.lbl_sweep_status.pack(fill="x", pady=(0, 8))

        tk.Frame(card, bg=BG_WHITE, height=4).pack()

    # ──────────────────────────────────────────────────────────
    #  DC BIAS CARD  (Phase 3)
    # ──────────────────────────────────────────────────────────
    def _build_bias_card(self, parent):
        card = self._white_card(parent, "DC Bias", "🔋",
                                fill="x", pady=(0, 8))

        # Voltage entry row
        row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=5)
        row.pack(fill="x")
        tk.Label(row, text="⚡", bg=BG_WHITE,
                fg=VIOLET, font=(SUI, 12), width=2).pack(side="left")
        tk.Label(row, text="Bias Voltage",
                bg=BG_WHITE, fg=T_DARK,
                font=F_BODY, width=16,
                anchor="w").pack(side="left", padx=(6, 0))
        ent_f = tk.Frame(row, bg="#EEF0FF", padx=8, pady=4)
        ent_f.pack(side="right")
        self.ent_bias_volt = tk.Entry(
            ent_f, textvariable=self.var_bias_volt,
            bg="#EEF0FF", fg=T_DARK,
            insertbackground=T_DARK,
            font=(SUI, 11, "bold"),
            relief="flat", width=8, bd=0)
        self.ent_bias_volt.pack(side="left")
        tk.Label(ent_f, text="V DC", bg="#EEF0FF",
                fg=T_SECOND, font=F_LABEL).pack(side="left", padx=(4, 0))
        tk.Frame(card, bg="#F0F0F5", height=1).pack(fill="x", padx=16)

        # Enable checkbox + Disable button row
        ctrl_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=8)
        ctrl_row.pack(fill="x")
        self.chk_bias_enable = tk.Checkbutton(
            ctrl_row, text="Enable Bias",
            variable=self.var_bias_enabled,
            command=self._toggle_bias,
            bg=BG_WHITE, fg=T_DARK,
            selectcolor="#EEF0FF",
            activebackground=BG_WHITE,
            font=F_BODY, cursor="hand2")
        self.chk_bias_enable.pack(side="left")
        self.btn_bias_disable = tk.Button(
            ctrl_row, text="Disable Bias",
            command=self._disable_bias,
            bg="#FFF0F0", fg=RED,
            font=F_SMALL, relief="flat",
            cursor="hand2", padx=10, pady=4, bd=0,
            state="disabled")
        self.btn_bias_disable.pack(side="right")
        tk.Frame(card, bg="#F0F0F5", height=1).pack(fill="x", padx=16)

        # Status row
        status_row = tk.Frame(card, bg=BG_WHITE, padx=16, pady=8)
        status_row.pack(fill="x")
        tk.Label(status_row, text="Status:",
                bg=BG_WHITE, fg=T_SECOND,
                font=F_SMALL).pack(side="left")
        self.lbl_bias_status = tk.Label(
            status_row, text="OFF",
            bg=BG_WHITE, fg=T_SECOND,
            font=(SUI, 10, "bold"))
        self.lbl_bias_status.pack(side="left", padx=(6, 0))

        tk.Frame(card, bg=BG_WHITE, height=4).pack()

    # ──────────────────────────────────────────────────────────
    #  INSTRUMENT INFO CARD
    # ──────────────────────────────────────────────────────────
    def _build_instrument_card(self, parent):
        card = self._white_card(parent, None, fill="x")

        body = tk.Frame(card, bg=BG_WHITE, padx=16, pady=12)
        body.pack(fill="x")

        left = tk.Frame(body, bg=BG_WHITE)
        left.pack(side="left")

        # Phone icon placeholder
        icon_c = tk.Canvas(left, width=44, height=52,
                           bg=BG_WHITE, highlightthickness=0)
        icon_c.pack(side="left", padx=(0, 14))
        icon_c.create_rectangle(4, 2, 40, 50, fill="#EEF0FF",
                                 outline="")
        icon_c.create_text(22, 26, text="📱", font=(SUI, 18))

        info = tk.Frame(body, bg=BG_WHITE)
        info.pack(side="left")
        tk.Label(info, text="Instrument Info",
                 bg=BG_WHITE, fg=T_SECOND,
                 font=F_SMALL).pack(anchor="w")
        tk.Label(info, text="Agilent Technologies",
                 bg=BG_WHITE, fg=T_DARK,
                 font=F_LABEL).pack(anchor="w")
        self.lbl_model = tk.Label(info, text="E4980A",
                 bg=BG_WHITE, fg=T_DARK,
                 font=(SUI, 13, "bold"))
        self.lbl_model.pack(anchor="w")

        right = tk.Frame(body, bg=BG_WHITE)
        right.pack(side="right")

        fw = tk.Frame(right, bg=BG_WHITE)
        fw.pack(side="left", padx=20)
        tk.Label(fw, text="Firmware",
                 bg=BG_WHITE, fg=T_SECOND,
                 font=F_TINY).pack(anchor="w")
        self.lbl_fw = tk.Label(fw, text="A.02.20",
                 bg=BG_WHITE, fg=T_DARK,
                 font=F_VALUE_SM)
        self.lbl_fw.pack(anchor="w")

        sn = tk.Frame(right, bg=BG_WHITE)
        sn.pack(side="left", padx=20)
        tk.Label(sn, text="S/N",
                 bg=BG_WHITE, fg=T_SECOND,
                 font=F_TINY).pack(anchor="w")
        self.lbl_sn = tk.Label(sn, text="MY46205236",
                 bg=BG_WHITE, fg=T_DARK,
                 font=F_VALUE_SM)
        self.lbl_sn.pack(anchor="w")

    # ──────────────────────────────────────────────────────────
    #  LIVE READINGS PANEL
    # ──────────────────────────────────────────────────────────
    def _build_live_readings(self, parent):
        # Outer card — gradient header effect with canvas
        outer = tk.Frame(parent, bg="#D0D0E0", padx=1, pady=1)
        outer.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        inner = tk.Frame(outer, bg=BG_WHITE)
        inner.pack(fill="both", expand=True)

        # Gradient header bar
        hdr_c = tk.Canvas(inner, height=44, bg=BG_WHITE,
                          highlightthickness=0)
        hdr_c.pack(fill="x")
        hdr_c.bind("<Configure>", lambda e: self._draw_grad_bar(hdr_c))

        hdr_overlay = tk.Frame(hdr_c, bg="")
        # Title over the gradient
        hdr_lbl = tk.Label(inner, text="",
                           bg=BG_WHITE)  # placeholder — drawn on canvas

        # Draw static gradient bar (approximated as violet→purple)
        def _draw_bar(c):
            c.delete("all")
            w = c.winfo_width() or 800
            # Simple 3-stop gradient approximation
            c.create_rectangle(0, 0, w, 44, fill=VIOLET, outline="")
            tk.Label(c, text="〜  LIVE READINGS",
                     bg=VIOLET, fg=T_WHITE,
                     font=F_SECTION,
                     padx=16, pady=12).place(x=0, y=0)
            # Mode badge right side
            self.lbl_mode_badge.place(x=w-120, y=8)
        hdr_c.bind("<Configure>", lambda e: _draw_bar(hdr_c))

        self.lbl_mode_badge = tk.Label(
            hdr_c, text="Mode: CPD",
            bg=BG_WHITE, fg=VIOLET,
            font=F_LABEL_B,
            padx=10, pady=4,
            relief="flat")

        # ── 3 mode reading rows ───────────────────────────────
        readings_area = tk.Frame(inner, bg=BG_WHITE)
        readings_area.pack(fill="x")
        readings_area.columnconfigure(0, weight=1)
        readings_area.columnconfigure(1, weight=0, minsize=2)
        readings_area.columnconfigure(2, weight=1)

        self._mode_rows = []
        self._reading_p1_labels = []
        self._reading_p2_labels = []

        for i in range(3):
            row_f = tk.Frame(readings_area, bg=BG_WHITE, pady=14, padx=20)
            row_f.grid(row=i, column=0, columnspan=3, sticky="ew")
            if i > 0:
                tk.Frame(readings_area, bg="#EBEBF0",
                         height=1).grid(row=i*2-1, column=0,
                                        columnspan=3, sticky="ew",
                                        padx=20)
            row_f.columnconfigure(0, weight=1)
            row_f.columnconfigure(1, weight=0, minsize=1)
            row_f.columnconfigure(2, weight=1)

            # P1 block
            p1f = tk.Frame(row_f, bg=BG_WHITE)
            p1f.grid(row=0, column=0, sticky="w")
            lbl_p1n_name = tk.Label(p1f, text="",
                                    bg=BG_WHITE,
                                    fg=MODE_COLORS[i],
                                    font=(SUI, 14, "bold"))
            lbl_p1n_name.pack(anchor="w")
            lbl_p1v = tk.Label(p1f, text="— — —",
                               bg=BG_WHITE,
                               fg=T_DARK if i == 0 else "#333",
                               font=F_VALUE_XL if i == 0 else F_VALUE_LG)
            lbl_p1v.pack(anchor="w")
            lbl_p1unit = tk.Label(p1f, text="",
                                  bg=BG_WHITE, fg=T_SECOND,
                                  font=(SUI, 12))
            lbl_p1unit.pack(anchor="w")

            # Divider
            tk.Frame(row_f, bg="#EBEBF0",
                     width=1).grid(row=0, column=1,
                                   sticky="ns", padx=30)

            # P2 block
            p2f = tk.Frame(row_f, bg=BG_WHITE)
            p2f.grid(row=0, column=2, sticky="w")
            lbl_p2n_name = tk.Label(p2f, text="",
                                    bg=BG_WHITE, fg=T_SECOND,
                                    font=(SUI, 14, "bold"))
            lbl_p2n_name.pack(anchor="w")
            lbl_p2v = tk.Label(p2f, text="— — —",
                               bg=BG_WHITE, fg=T_DARK,
                               font=F_VALUE_XL if i == 0 else F_VALUE_LG)
            lbl_p2v.pack(anchor="w")

            # Status badge
            lbl_st = tk.Label(row_f, text="",
                              bg="#E8FFF0", fg=GREEN,
                              font=F_LABEL_B,
                              padx=8, pady=2)
            lbl_st.grid(row=0, column=2, sticky="ne")

            # Mode badge
            lbl_m = tk.Label(row_f, text="",
                             bg=BG_WHITE,
                             fg=MODE_COLORS[i],
                             font=(SUI, 8, "bold"))
            lbl_m.grid(row=0, column=0, sticky="ne")

            self._mode_rows.append(
                (lbl_m, lbl_p1n_name, lbl_p1v, lbl_p2n_name, lbl_p2v, lbl_st))
            self._reading_p1_labels.append((lbl_p1v, lbl_p1unit))

        # ── Status pill row ───────────────────────────────────
        status_row = tk.Frame(inner, bg="#F8F9FF", pady=10)
        status_row.pack(fill="x")

        def status_pill(icon, label, val_init, color=T_DARK):
            pill = tk.Frame(status_row, bg="#F8F9FF", padx=14)
            pill.pack(side="left", fill="x", expand=True)
            hd = tk.Frame(pill, bg="#F8F9FF")
            hd.pack(anchor="w")
            tk.Label(hd, text=icon, bg="#F8F9FF",
                     fg=VIOLET, font=(SUI, 11)).pack(side="left")
            tk.Label(hd, text=f"  {label}",
                     bg="#F8F9FF", fg=T_SECOND,
                     font=F_TINY).pack(side="left")
            lv = tk.Label(pill, text=val_init,
                          bg="#F8F9FF", fg=color,
                          font=F_VALUE_SM)
            lv.pack(anchor="w")
            return lv

        self.lbl_freq_disp  = status_pill("〜", "Frequency",  "100.000 kHz", VIOLET)
        self.lbl_volt_disp  = status_pill("⚡", "Voltage",    "1.000 V",     ORANGE)
        self.lbl_aper_disp  = status_pill("🕐", "Aperture",   "MED")
        self.lbl_bias_disp  = status_pill("🔋", "Bias",       "0.000 V",     T_DARK)
        self.lbl_status_pill= status_pill("✅", "Status",     "STABLE",      GREEN)

        # ── START / STOP + counters ───────────────────────────
        ctrl_row = tk.Frame(inner, bg=BG_WHITE, pady=10, padx=16)
        ctrl_row.pack(fill="x")

        self.btn_start = tk.Button(
            ctrl_row, text="▶  START",
            command=self._start,
            bg=GREEN, fg=T_WHITE,
            font=F_BTN_LG, relief="flat",
            activebackground="#25a244",
            cursor="hand2",
            padx=28, pady=14, bd=0)
        self.btn_start.pack(side="left", padx=(0, 10))

        self.btn_pause = tk.Button(
            ctrl_row, text="⏸  PAUSE",
            command=self._toggle_pause,
            bg=AMBER, fg=T_DARK,
            font=F_BTN_LG, relief="flat",
            activebackground="#e0bf09",
            cursor="hand2",
            padx=22, pady=14, bd=0,
            state="disabled")
        self.btn_pause.pack(side="left", padx=(0, 10))

        self.btn_stop = tk.Button(
            ctrl_row, text="■  STOP",
            command=self._stop,
            bg=RED, fg=T_WHITE,
            font=F_BTN_LG, relief="flat",
            activebackground="#c93a30",
            cursor="hand2",
            padx=28, pady=14, bd=0,
            state="disabled")
        self.btn_stop.pack(side="left", padx=(0, 20))

        # Reading counter
        ctr_f = tk.Frame(ctrl_row, bg=BG_WHITE)
        ctr_f.pack(side="left", padx=14)
        tk.Label(ctr_f, text="Reading",
                 bg=BG_WHITE, fg=T_SECOND,
                 font=F_TINY).pack(anchor="w")
        self.lbl_n_disp = tk.Label(ctr_f, text="0 / 100",
                                    bg=BG_WHITE, fg=T_DARK,
                                    font=F_VALUE_MD)
        self.lbl_n_disp.pack(anchor="w")

        # Progress ring (text-based)
        prog_f = tk.Frame(ctrl_row, bg=BG_WHITE)
        prog_f.pack(side="left", padx=6)
        self.lbl_prog = tk.Label(prog_f, text="0%",
                                  bg="#EEF0FF", fg=VIOLET,
                                  font=(SUI, 11, "bold"),
                                  padx=12, pady=8)
        self.lbl_prog.pack()

        # Elapsed
        elap_f = tk.Frame(ctrl_row, bg=BG_WHITE, padx=14)
        elap_f.pack(side="left")
        tk.Label(elap_f, text="⏱️  Elapsed Time",
                 bg=BG_WHITE, fg=T_SECOND,
                 font=F_TINY).pack(anchor="w")
        self.lbl_elapsed = tk.Label(elap_f, text="00:00:00",
                                     bg=BG_WHITE, fg=T_DARK,
                                     font=F_VALUE_MD)
        self.lbl_elapsed.pack(anchor="w")

        # ── Visual progress bar (full width, below controls) ──
        bar_row = tk.Frame(inner, bg=BG_WHITE, padx=16)
        bar_row.pack(fill="x", pady=(0, 12))
        style = ttk.Style()
        style.configure("LCR.Horizontal.TProgressbar",
                        troughcolor="#EEF0FF",
                        bordercolor="#EEF0FF",
                        background=VIOLET,
                        lightcolor=VIOLET,
                        darkcolor=VIOLET,
                        thickness=10)
        self.progressbar = ttk.Progressbar(
            bar_row, style="LCR.Horizontal.TProgressbar",
            orient="horizontal", mode="determinate",
            maximum=100, value=0)
        self.progressbar.pack(fill="x")

    # ──────────────────────────────────────────────────────────
    #  GRAPH AREA
    # ──────────────────────────────────────────────────────────
    def _build_graph_area(self, parent):
        outer = tk.Frame(parent, bg="#D0D0E0", padx=1, pady=1)
        outer.grid(row=1, column=0, sticky="nsew")
        self.graph_card_outer = outer   # used by sidebar "Graph" nav scroll

        inner = tk.Frame(outer, bg=BG_CARD)
        inner.pack(fill="both", expand=True)
        inner.rowconfigure(1, weight=1)
        inner.columnconfigure(0, weight=1)

        # Graph header
        ghdr = tk.Frame(inner, bg=GRAD_MID, pady=10, padx=16)
        ghdr.grid(row=0, column=0, sticky="ew")
        tk.Label(ghdr, text="📈  LIVE GRAPH",
                 bg=GRAD_MID, fg=T_WHITE,
                 font=F_SECTION).pack(side="left")

        # Graph mode selector — choose WHICH active mode to plot
        sel_f = tk.Frame(ghdr, bg=GRAD_MID)
        sel_f.pack(side="left", padx=(18, 0))
        tk.Label(sel_f, text="Graphing:",
                 bg=GRAD_MID, fg=T_LIGHT,
                 font=F_SMALL).pack(side="left", padx=(0, 6))
        self.var_graph_mode = tk.StringVar(value="")
        self.cb_graph_mode = ttk.Combobox(
            sel_f, textvariable=self.var_graph_mode,
            state="readonly", width=10,
            font=F_SMALL, values=[])
        self.cb_graph_mode.pack(side="left")
        self.cb_graph_mode.bind(
            "<<ComboboxSelected>>", self._on_graph_mode_selected)

        # Legend / status caption (right)
        leg_f = tk.Frame(ghdr, bg=GRAD_MID)
        leg_f.pack(side="right")
        self.lbl_legend = tk.Label(leg_f, text="— P1 (cyan)   — P2 (pink)",
                                    bg=GRAD_MID, fg=T_WHITE,
                                    font=F_SMALL)
        self.lbl_legend.pack()

        # Stats row below graph header
        stat_row = tk.Frame(inner, bg=BG_CARD, pady=6, padx=12)
        stat_row.grid(row=1, column=0, sticky="ew")

        self._stat_rows = []
        for i in range(3):
            sf = tk.Frame(stat_row, bg="#1E1E30",
                          padx=10, pady=6)
            sf.pack(side="left", padx=4, fill="y")
            lm = tk.Label(sf, text="",
                          bg="#1E1E30", fg=MODE_COLORS[i],
                          font=(SUI, 8, "bold"))
            lm.pack(anchor="w")
            lmin  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_SECOND, font=F_MONO_SM)
            lmax  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_SECOND, font=F_MONO_SM)
            lmean = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_WHITE,  font=F_MONO_SM)
            lstd  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_DIM,   font=F_MONO_SM)
            for lb in [lmin, lmax, lmean, lstd]:
                lb.pack(anchor="w")
            self._stat_rows.append((lm, lmin, lmax, lmean, lstd))

        # Graph canvas — dual Matplotlib graphs (P1 / P2 vs Time)
        self.graph = DualLiveGraph(inner, max_points=GRAPH_POINTS)
        self.graph.grid(row=2, column=0, sticky="nsew",
                        padx=8, pady=8)
        inner.rowconfigure(2, weight=1)

        # Log strip
        log_outer = tk.Frame(inner, bg=BG_CARD)
        log_outer.grid(row=3, column=0, sticky="ew")
        log_hdr = tk.Frame(log_outer, bg="#1E1E30", padx=12, pady=4)
        log_hdr.pack(fill="x")
        tk.Label(log_hdr, text="💬  Data Logging",
                 bg="#1E1E30", fg=T_SECOND,
                 font=F_SMALL).pack(side="left")
        self.lbl_log_on = tk.Label(log_hdr, text="● ON",
                                    bg="#1E1E30", fg=GREEN,
                                    font=F_LABEL_B)
        self.lbl_log_on.pack(side="left", padx=8)
        self.lbl_autosave = tk.Label(log_hdr,
                                      text="↺  Auto Save ✓",
                                      bg="#1E1E30", fg=T_SECOND,
                                      font=F_SMALL)
        self.lbl_autosave.pack(side="right")

        self.txt_log = tk.Text(
            log_outer, height=3,
            bg="#0A0A14", fg="#7A8A9A",
            font=F_MONO_SM, bd=0,
            state="disabled", wrap="word",
            insertbackground=T_WHITE)
        self.txt_log.pack(fill="x", padx=8, pady=(0, 8))

    # ──────────────────────────────────────────────────────────
    #  FREQUENCY SWEEP GRAPH CARD  (Phase 2)
    # ──────────────────────────────────────────────────────────
    def _build_sweep_graph_area(self, parent):
        outer = tk.Frame(parent, bg="#D0D0E0", padx=1, pady=1)
        outer.grid(row=2, column=0, sticky="nsew", pady=(8, 0))
        self.sweep_card_outer = outer

        inner = tk.Frame(outer, bg=BG_CARD)
        inner.pack(fill="both", expand=True)
        inner.rowconfigure(1, weight=1)
        inner.columnconfigure(0, weight=1)

        hdr = tk.Frame(inner, bg=GRAD_MID, pady=10, padx=16)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="📡  FREQUENCY SWEEP (BODE PLOT)",
                bg=GRAD_MID, fg=T_WHITE,
                font=F_SECTION).pack(side="left")

        self.sweep_graph = BodeSweepGraph(inner)
        self.sweep_graph.grid(row=1, column=0, sticky="nsew",
                              padx=8, pady=8)
        inner.rowconfigure(1, weight=1, minsize=320)

    # ──────────────────────────────────────────────────────────
    #  DATA LOGGER / STATISTICS CARD  (new)
    # ──────────────────────────────────────────────────────────
    def _build_datalogger_card(self, parent):
        outer = tk.Frame(parent, bg="#D0D0E0", padx=1, pady=1)
        outer.grid(row=1, column=0, columnspan=2, sticky="nsew",
                   pady=(10, 0))
        self.datalogger_card_outer = outer

        inner = tk.Frame(outer, bg=BG_CARD)
        inner.pack(fill="both", expand=True)
        inner.columnconfigure(0, weight=1)
        inner.columnconfigure(1, weight=1)

        hdr = tk.Frame(inner, bg=GRAD_MID, pady=10, padx=16)
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        tk.Label(hdr, text="📋  DATA LOGGER — STATISTICS & SESSION HISTORY",
                bg=GRAD_MID, fg=T_WHITE, font=F_SECTION).pack(side="left")

        # ── A. Live scrollable readings table ──────────────────
        tbl_outer = tk.Frame(inner, bg=BG_CARD, padx=10, pady=10)
        tbl_outer.grid(row=1, column=0, sticky="nsew")
        inner.rowconfigure(1, weight=1)

        tk.Label(tbl_outer, text="Live Readings (last 500)",
                bg=BG_CARD, fg=T_SECOND, font=F_LABEL_B
                ).pack(anchor="w", pady=(0, 4))

        cols = ("n", "time", "mode", "p1", "p2", "status")
        headings = {"n":"#", "time":"Time", "mode":"Mode",
                    "p1":"P1", "p2":"P2", "status":"Status"}
        widths = {"n":40, "time":90, "mode":60,
                  "p1":110, "p2":110, "status":80}

        style = ttk.Style()
        style.configure("DL.Treeview",
                        background="#0A0A14", fieldbackground="#0A0A14",
                        foreground=T_LIGHT, rowheight=22,
                        font=F_MONO_SM, borderwidth=0)
        style.configure("DL.Treeview.Heading",
                        background="#1E1E30", foreground=T_SECOND,
                        font=F_SMALL)
        style.map("DL.Treeview", background=[("selected", GRAD_MID)])

        tree_frame = tk.Frame(tbl_outer, bg=BG_CARD)
        tree_frame.pack(fill="both", expand=True)
        self.dl_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            style="DL.Treeview", height=12)
        for c in cols:
            self.dl_tree.heading(c, text=headings[c])
            self.dl_tree.column(c, width=widths[c], anchor="center")
        self.dl_tree.tag_configure("ok", foreground=GREEN)
        self.dl_tree.tag_configure("bad", foreground=RED)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self.dl_tree.yview)
        self.dl_tree.configure(yscrollcommand=vsb.set)
        self.dl_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # ── B. Statistics summary cards ─────────────────────────
        stats_outer = tk.Frame(inner, bg=BG_CARD, padx=10, pady=10)
        stats_outer.grid(row=1, column=1, sticky="nsew")
        tk.Label(stats_outer, text="Live Statistics (this run)",
                bg=BG_CARD, fg=T_SECOND, font=F_LABEL_B
                ).pack(anchor="w", pady=(0, 4))

        self._dl_stat_rows = []
        stat_grid = tk.Frame(stats_outer, bg=BG_CARD)
        stat_grid.pack(fill="x")
        for i in range(3):
            sf = tk.Frame(stat_grid, bg="#1E1E30", padx=10, pady=6)
            sf.pack(side="left", padx=4, fill="both", expand=True)
            lm = tk.Label(sf, text="", bg="#1E1E30",
                         fg=MODE_COLORS[i], font=(SUI, 9, "bold"))
            lm.pack(anchor="w")
            lcount = tk.Label(sf, text="", bg="#1E1E30",
                              fg=T_DIM, font=F_MONO_SM)
            lmin  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_SECOND, font=F_MONO_SM)
            lmax  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_SECOND, font=F_MONO_SM)
            lmean = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_WHITE,  font=F_MONO_SM)
            lstd  = tk.Label(sf, text="", bg="#1E1E30",
                             fg=T_DIM,   font=F_MONO_SM)
            for lb in [lcount, lmin, lmax, lmean, lstd]:
                lb.pack(anchor="w")
            self._dl_stat_rows.append(
                (lm, lcount, lmin, lmax, lmean, lstd))

        # ── C. Histogram / distribution chart ───────────────────
        hist_hdr = tk.Frame(stats_outer, bg=BG_CARD)
        hist_hdr.pack(fill="x", pady=(10, 2))
        tk.Label(hist_hdr, text="Distribution (P1)",
                bg=BG_CARD, fg=T_SECOND, font=F_LABEL_B
                ).pack(side="left")
        self.cb_hist_mode = ttk.Combobox(
            hist_hdr, textvariable=self._dl_hist_mode,
            state="readonly", width=8, font=F_SMALL, values=[])
        self.cb_hist_mode.pack(side="right")
        self.cb_hist_mode.bind(
            "<<ComboboxSelected>>",
            lambda e: self._draw_histogram(self._dl_hist_mode.get()))

        self._hist_fig = Figure(figsize=(4.2, 2.1), dpi=90,
                                facecolor=BG_GRAPH)
        self._hist_ax = self._hist_fig.add_subplot(111)
        self._hist_canvas = FigureCanvasTkAgg(
            self._hist_fig, master=stats_outer)
        self._hist_canvas.get_tk_widget().pack(
            fill="both", expand=True, pady=(4, 0))
        self._draw_histogram(None)

        # ── D. Session history ──────────────────────────────────
        hist_outer = tk.Frame(inner, bg=BG_CARD, padx=10, pady=10)
        hist_outer.grid(row=2, column=0, columnspan=2, sticky="ew")
        tk.Label(hist_outer, text="Session History — click a run to reload its statistics",
                bg=BG_CARD, fg=T_SECOND, font=F_LABEL_B
                ).pack(anchor="w", pady=(0, 4))

        lb_frame = tk.Frame(hist_outer, bg=BG_CARD)
        lb_frame.pack(fill="x")
        self.lb_sessions = tk.Listbox(
            lb_frame, height=5, bg="#0A0A14", fg=T_LIGHT,
            font=F_MONO_SM, bd=0, highlightthickness=0,
            selectbackground=GRAD_MID, activestyle="none")
        self.lb_sessions.pack(side="left", fill="both", expand=True)
        lb_vsb = ttk.Scrollbar(lb_frame, orient="vertical",
                               command=self.lb_sessions.yview)
        self.lb_sessions.configure(yscrollcommand=lb_vsb.set)
        lb_vsb.pack(side="right", fill="y")
        self.lb_sessions.bind("<<ListboxSelect>>",
                              self._on_session_selected)

    # ──────────────────────────────────────────────────────────
    #  DATA LOGGER — LIVE REFRESH  (called each measurement cycle)
    # ──────────────────────────────────────────────────────────
    def _refresh_datalogger(self, results, active_modes, aperture):
        # Rebuild the table from the rolling readings_log — cheap
        # (<=500 rows) and avoids any risk of the tree drifting out
        # of sync with the underlying data.
        self.dl_tree.delete(*self.dl_tree.get_children())
        for row in self.readings_log:
            tag = "ok" if row["status"] == "OK" else "bad"
            self.dl_tree.insert(
                "", "end",
                values=(row["n"], row["time"], row["mode"],
                        row["p1"], row["p2"], row["status"]),
                tags=(tag,))
        if self.dl_tree.get_children():
            self.dl_tree.see(self.dl_tree.get_children()[-1])

        # Stats cards
        for i in range(3):
            lm, lcount, lmin, lmax, lmean, lstd = self._dl_stat_rows[i]
            if i < len(active_modes):
                mode = active_modes[i]
                vals = self._stats_data.get(mode, [])
                pt1, _ = MODE_PARAM_TYPES.get(scpi_code(mode), ("R", "X"))
                lm.config(text=mode)
                lcount.config(text=f"n    {len(vals)}")
                if vals:
                    lmin.config(text=f"min  {format_lcr_value(min(vals), pt1, aperture)}")
                    lmax.config(text=f"max  {format_lcr_value(max(vals), pt1, aperture)}")
                    lmean.config(text=f"avg  {format_lcr_value(sum(vals)/len(vals), pt1, aperture)}")
                    if len(vals) > 1:
                        lstd.config(text=f"std  {format_lcr_value(_stats.stdev(vals), pt1, aperture)}")
                    else:
                        lstd.config(text="")
            else:
                for lb in [lm, lcount, lmin, lmax, lmean, lstd]:
                    lb.config(text="")

        # Histogram mode dropdown — keep in sync with active modes
        self.cb_hist_mode["values"] = active_modes
        if active_modes and self._dl_hist_mode.get() not in active_modes:
            self._dl_hist_mode.set(active_modes[0])
        self._draw_histogram(self._dl_hist_mode.get())

    def _draw_histogram(self, mode):
        self._hist_ax.clear()
        self._hist_ax.set_facecolor(BG_GRAPH)
        vals = self._stats_data.get(mode, []) if mode else []
        if vals and len(vals) > 1:
            n_bins = min(20, max(5, len(vals) // 3))
            self._hist_ax.hist(vals, bins=n_bins, color=CYAN,
                               edgecolor=BG_GRAPH, alpha=0.85)
            if len(vals) > 2 and _stats.stdev(vals) > 0:
                mu, sigma = _stats.mean(vals), _stats.stdev(vals)
                xs = [mu + sigma * (k - 100) / 25 for k in range(201)]
                # Scale the normal curve to roughly match bar heights
                bin_width = (max(vals) - min(vals)) / n_bins or 1
                scale = len(vals) * bin_width
                ys = [scale * (1 / (sigma * math.sqrt(2*math.pi))) *
                     math.exp(-0.5 * ((x - mu) / sigma) ** 2) for x in xs]
                self._hist_ax.plot(xs, ys, color=PINK, linewidth=1.4)
        self._hist_ax.tick_params(colors=T_SECOND, labelsize=7)
        for spine in self._hist_ax.spines.values():
            spine.set_color(G_GRID)
        self._hist_fig.tight_layout(pad=0.4)
        self._hist_canvas.draw_idle()

    # ──────────────────────────────────────────────────────────
    #  SESSION HISTORY — click to reload
    # ──────────────────────────────────────────────────────────
    def _record_session(self):
        """Called from _finish() once a run completes — snapshots
        this run's stats into session_history and refreshes the list."""
        if self.reading_count == 0:
            return
        active_modes = list(self._stats_data.keys())
        per_mode = {}
        for mode in active_modes:
            vals = self._stats_data.get(mode, [])
            if not vals:
                continue
            per_mode[mode] = {
                "count": len(vals),
                "min": min(vals), "max": max(vals),
                "mean": sum(vals) / len(vals),
                "std": _stats.stdev(vals) if len(vals) > 1 else 0.0,
                "values": list(vals),   # kept for histogram reload
            }
        entry = {
            "start_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "modes": active_modes,
            "freq": self.var_freq_hz.get(),
            "volt": self.var_volt_var.get(),
            "aperture": self.var_aperture.get(),
            "average": self.var_average.get(),
            "n_readings": self.reading_count,
            "per_mode": per_mode,
        }
        self.session_history.append(entry)
        label = (f"{entry['start_time']}   {'+'.join(active_modes) or '—'}   "
                f"{_hz_label(entry['freq'])}   {entry['n_readings']} rdgs")
        self.lb_sessions.insert("end", label)

    def _on_session_selected(self, event=None):
        sel = self.lb_sessions.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx >= len(self.session_history):
            return
        self._selected_session_idx = idx
        entry = self.session_history[idx]
        aperture = entry.get("aperture", "MED")
        modes = entry["modes"]
        for i in range(3):
            lm, lcount, lmin, lmax, lmean, lstd = self._dl_stat_rows[i]
            if i < len(modes) and modes[i] in entry["per_mode"]:
                mode = modes[i]
                s = entry["per_mode"][mode]
                pt1, _ = MODE_PARAM_TYPES.get(scpi_code(mode), ("R", "X"))
                lm.config(text=f"{mode} (saved)")
                lcount.config(text=f"n    {s['count']}")
                lmin.config(text=f"min  {format_lcr_value(s['min'], pt1, aperture)}")
                lmax.config(text=f"max  {format_lcr_value(s['max'], pt1, aperture)}")
                lmean.config(text=f"avg  {format_lcr_value(s['mean'], pt1, aperture)}")
                lstd.config(text=f"std  {format_lcr_value(s['std'], pt1, aperture)}")
            else:
                for lb in [lm, lcount, lmin, lmax, lmean, lstd]:
                    lb.config(text="")
        self.cb_hist_mode["values"] = modes
        if modes:
            self._dl_hist_mode.set(modes[0])
            self._draw_saved_histogram(entry, modes[0])

    def _draw_saved_histogram(self, entry, mode):
        vals = entry["per_mode"].get(mode, {}).get("values", [])
        self._stats_data_backup = self._stats_data
        self._stats_data = {mode: vals}
        self._draw_histogram(mode)
        self._stats_data = self._stats_data_backup

    # ──────────────────────────────────────────────────────────
    #  AUTO-SAVE / RELOAD SESSION  (Feature 2)
    # ──────────────────────────────────────────────────────────
    def _autosave_tick(self):
        """Runs every 60s regardless of measurement state — writes
        the current settings + readings + stats + session history
        to a small JSON file so the app can offer to resume on the
        next launch."""
        try:
            self._save_session_to_disk()
        except Exception as e:
            self.log(f"Autosave failed: {e}")
        self.root.after(60000, self._autosave_tick)

    def _save_session_to_disk(self):
        data = {
            "settings": {
                "freq": self.var_freq_hz.get(),
                "volt": self.var_volt_var.get(),
                "aperture": self.var_aperture.get(),
                "average": self.var_average.get(),
                "n": self.var_n.get(),
                "interval": self.var_interval.get(),
                "active_modes": [MODES[j] for j, v in
                                enumerate(self.selected_modes) if v.get()],
            },
            "readings": list(self.readings_log),
            "stats": {m: v for m, v in self._stats_data.items()},
            "session_history": self.session_history,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        with open(self._autosave_path, "w", encoding="utf-8") as f:
            json.dump(data, f)

    def _offer_session_reload(self):
        if not os.path.exists(self._autosave_path):
            return
        try:
            with open(self._autosave_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return
        if not data.get("readings") and not data.get("session_history"):
            return
        saved_at = data.get("saved_at", "an earlier session")
        if messagebox.askyesno(
                "Reload previous session?",
                f"An autosaved session from {saved_at} was found "
                f"({len(data.get('readings', []))} readings, "
                f"{len(data.get('session_history', []))} past runs).\n\n"
                "Reload it into the Data Logger tab?"):
            self._reload_session(data)

    def _reload_session(self, data):
        self.readings_log = deque(data.get("readings", []),
                                  maxlen=DATALOGGER_MAX_ROWS)
        self._stats_data = data.get("stats", {})
        self.session_history = data.get("session_history", [])
        for entry in self.session_history:
            active_modes = entry["modes"]
            label = (f"{entry['start_time']}   {'+'.join(active_modes) or '—'}   "
                    f"{_hz_label(entry['freq'])}   {entry['n_readings']} rdgs")
            self.lb_sessions.insert("end", label)
        aperture = data.get("settings", {}).get("aperture", "MED")
        active_modes = list(self._stats_data.keys())
        self.dl_tree.delete(*self.dl_tree.get_children())
        for row in self.readings_log:
            tag = "ok" if row["status"] == "OK" else "bad"
            self.dl_tree.insert(
                "", "end",
                values=(row["n"], row["time"], row["mode"],
                        row["p1"], row["p2"], row["status"]),
                tags=(tag,))
        self.cb_hist_mode["values"] = active_modes
        if active_modes:
            self._dl_hist_mode.set(active_modes[0])
            self._draw_histogram(active_modes[0])
        self.log("Reloaded autosaved Data Logger session.")

    # ──────────────────────────────────────────────────────────
    #  MODE BUTTON TOGGLE STYLE
    # ──────────────────────────────────────────────────────────
    def _refresh_mode_styles(self):
        active = [j for j, v in enumerate(self.selected_modes) if v.get()]
        for i, btn in enumerate(self._mode_btns):
            base_bg = MODE_CARD_COLORS[i % len(MODE_CARD_COLORS)]
            base_fg = MODE_CARD_FG[i % len(MODE_CARD_FG)]
            if self.selected_modes[i].get():
                rank = active.index(i)
                outline_col = MODE_COLORS[rank % len(MODE_COLORS)]
                btn.config(
                    bg=base_bg,
                    fg=base_fg,
                    font=(SUI, 9, "bold"),
                    relief="solid",
                    bd=2,
                    highlightbackground=outline_col,
                    highlightthickness=2,
                    highlightcolor=outline_col)
                # Simulate check mark
                btn.config(text=f"✓ {MODES[i]}")
            else:
                btn.config(
                    bg=base_bg, fg=base_fg,
                    font=(SUI, 9),
                    relief="flat", bd=0,
                    highlightthickness=0)
                btn.config(text=MODES[i])

    def _toggle_mode(self, i):
        active = [j for j, v in enumerate(self.selected_modes) if v.get()]
        if self.selected_modes[i].get():
            self.selected_modes[i].set(False)
        else:
            if len(active) >= 3:
                self.lbl_mode_warn.config(
                    text="⚠  Max 3 modes allowed")
                return
            self.selected_modes[i].set(True)
        self.lbl_mode_warn.config(text="")
        self._refresh_mode_styles()
        # Update mode badge
        active_names = [MODES[j]
                        for j, v in enumerate(self.selected_modes)
                        if v.get()]
        self.lbl_mode_badge.config(
            text="Mode: " + ", ".join(active_names) if active_names else "")
        self._refresh_graph_mode_options(active_names)

    # ──────────────────────────────────────────────────────────
    #  GRAPH MODE SELECTOR  (which active mode to plot)
    # ──────────────────────────────────────────────────────────
    def _refresh_graph_mode_options(self, active_names):
        """Keep the 'Graphing:' dropdown in sync with the currently
        ticked measurement modes. Preserves the current selection
        if it is still valid; otherwise defaults to the first mode."""
        self.cb_graph_mode["values"] = active_names
        current = self.var_graph_mode.get()
        if active_names:
            if current not in active_names:
                self.var_graph_mode.set(active_names[0])
            self.graph.set_selected_mode(self.var_graph_mode.get())
        else:
            self.var_graph_mode.set("")

    def _on_graph_mode_selected(self, event=None):
        mode = self.var_graph_mode.get()
        if mode:
            self.graph.set_selected_mode(mode)
            self.graph.redraw(
                [MODES[j] for j, v in enumerate(self.selected_modes)
                 if v.get()],
                self.var_aperture.get())

    # ──────────────────────────────────────────────────────────
    #  FREQUENCY VALIDATION  (preserved)

    # ──────────────────────────────────────────────────────────
    def _validate_freq(self, *_):
        try:
            hz = float(self.var_freq_hz.get())
            if hz < 20:
                self.lbl_freq_hint.config(text="⚠ min 20 Hz", fg=RED)
            elif hz > 2_000_000:
                self.lbl_freq_hint.config(text="⚠ max 2 MHz", fg=RED)
            else:
                if hz >= 1_000_000:
                    hint = f"= {hz/1e6:.4g} MHz"
                elif hz >= 1000:
                    hint = f"= {hz/1e3:.4g} kHz"
                else:
                    hint = f"= {hz:.0f} Hz"
                self.lbl_freq_hint.config(text=hint, fg=GREEN)
        except ValueError:
            self.lbl_freq_hint.config(text="⚠ numbers only", fg=RED)

    # ──────────────────────────────────────────────────────────
    #  CONNECT  (preserved)
    # ──────────────────────────────────────────────────────────
    # ──────────────────────────────────────────────────────────
    #  INSTRUMENT BUSY GUARD
    #  GPIB only supports one command/response transaction at a
    #  time. Measurement, Calibration, Frequency Sweep, and DC
    #  Bias all talk to the instrument, so only one of them may
    #  run at once — running two concurrently would interleave
    #  GPIB traffic and corrupt readings or hang the bus.
    # ──────────────────────────────────────────────────────────
    def _instrument_busy(self):
        """Returns a human-readable reason the instrument is busy,
        or None if it's free for a new operation to claim."""
        if self.running:
            return "a measurement is currently running"
        if self.calibrating:
            return "a calibration is currently running"
        if self.sweeping:
            return "a frequency sweep is currently running"
        if self.bias_busy:
            return "a DC bias command is in progress"
        return None

    def _connect(self):
        try:
            rm = pyvisa.ResourceManager()
            self.lcr = rm.open_resource(GPIB_ADDRESS)
            self.lcr.read_termination  = "\n"
            self.lcr.write_termination = "\n"
            self.lcr.timeout = 10000
            self.lcr.write("*CLS")
            # Force plain ASCII (comma-separated) response data.
            # If the instrument is left in REAL (binary IEEE-754
            # block) format from a prior session, every FETCh?
            # response starts with a "#<len digits><length>" header
            # followed by raw binary bytes -- pyvisa's query() tries
            # to decode that as ASCII text and dies with
            # "'ascii' codec can't decode byte 0x.. at position N"
            # at a fixed byte offset every time (the header length).
            # This must be sent before any FETCh?/MEAS?/READ? call.
            self.lcr.write("FORMAT:DATA ASCII")
            idn = self.lcr.query("*IDN?").strip()
            self._idn_str = idn
            print("IDN =", idn)

            self.lcr.write("FUNC:IMP VDCIDC")
            print("After VDCIDC =", self.lcr.query("FUNC:IMP?").strip())

            self.lcr.write("FUNC:IMP LPRDC")
            print("After LPRDC =", self.lcr.query("FUNC:IMP?").strip())

            self.lcr.write("FUNC:IMP LSRDC")
            print("After LSRDC =", self.lcr.query("FUNC:IMP?").strip())
        

            self.lbl_conn_dot.config(fg=GREEN)
            self.lbl_conn_text.config(
                text="  Connected", fg=GREEN)
            self.btn_connect.config(
                text="✓  Connected\n(reconnect)",
                bg=GREEN)
            self.log(f"Connected: {idn}")

            # Parse IDN for instrument card
            parts = idn.split(",")
            if len(parts) >= 4:
                self.lbl_model.config(text=parts[1].strip())
                self.lbl_fw.config(text=parts[3].strip())
                self.lbl_sn.config(text=parts[2].strip())

        except Exception as e:
            messagebox.showerror("Connection Error", str(e))
            self.log(f"ERROR: {e}")

    # ──────────────────────────────────────────────────────────
    #  OPEN / SHORT CALIBRATION
    #  Uses the E4980A's own correction registers (CORR:OPEN/SHORT)
    #  rather than manual Python subtraction — once enabled, every
    #  FETCh? from the instrument is already compensated, so the
    #  rest of the app needs no changes to benefit from it.
    # ──────────────────────────────────────────────────────────
    def _run_open_cal(self):
        self._run_calibration(
            kind="OPEN",
            instruction=(
                "OPEN CIRCUIT CALIBRATION\n\n"
                "Disconnect or open the test fixture / DUT leads "
                "completely (nothing connected).\n\n"
                "Click OK to measure the open-circuit reference."),
            scpi_exec="CORR:OPEN:EXEC",
            scpi_enable="CORR:OPEN ON",
            button=lambda: self.btn_open_cal,
            label_base="Open Cal",
            done_attr="open_cal_done")
        self.save_calibration()

    def _run_short_cal(self):
        self._run_calibration(
            kind="SHORT",
            instruction=(
                "SHORT CIRCUIT CALIBRATION\n\n"
                "Connect a short (shorting bar / direct short) "
                "across the test fixture leads.\n\n"
                "Click OK to measure the short-circuit reference."),
            scpi_exec="CORR:SHORT:EXEC",
            scpi_enable="CORR:SHORT ON",
            button=lambda: self.btn_short_cal,
            label_base="Short Cal",
            done_attr="short_cal_done")
        self.save_calibration()

    def _drain_stale_response(self):
        """Discard any unread bytes sitting in the instrument's output
        queue from a previous query whose response we never fully
        read (e.g. a query that timed out on our end while the
        instrument was still busy). Sending a *new* query while an
        old, unread response is still pending is exactly what
        triggers a -420 'Query UNTERMINATED' error on the instrument
        — which then poisons every later query in the session until
        it's cleared."""
        if self.lcr is None:
            return
        try:
            old_timeout = self.lcr.timeout
            self.lcr.timeout = 200
            while True:
                self.lcr.read_raw()
        except Exception:
            pass
        finally:
            try:
                self.lcr.timeout = old_timeout
            except Exception:
                pass

    def _run_calibration(self, kind, instruction, scpi_exec,
                         scpi_enable, button, label_base, done_attr):
        if self.lcr is None:
            messagebox.showwarning(
                "Not Connected",
                "Connect to the instrument before running calibration.")
            return
        busy = self._instrument_busy()
        if busy:
            messagebox.showwarning(
                "Instrument Busy",
                f"Cannot calibrate — {busy}. Stop it first.")
            return

        proceed = messagebox.askokcancel(
            f"{kind} Calibration", instruction)
        if not proceed:
            return

        self.calibrating = True
        btn = button()
        btn.config(text=f"{label_base}  ⏳", state="disabled")
        self.lbl_cal_status.config(
            text=f"Running {kind.lower()} calibration — please wait...",
            fg=AMBER)

        def worker():
            # How long we're willing to wait in total for the
            # correction sweep to finish, and how often we check in.
            CAL_MAX_WAIT_S = 120
            POLL_INTERVAL_S = 1.0
            try:
                # Flush any stale bytes sitting in the GPIB buffers
                # from a previous command before we start — a leftover
                # byte here is a common cause of an immediate/early
                # VI_ERROR_TMO on the very next read.
                try:
                    self.lcr.clear()
                except Exception:
                    pass
                self._drain_stale_response()

                self.lcr.write("*CLS")     # clear status registers
                self.lcr.write("*ESE 1")   # enable OPC bit to propagate to STB's ESB bit —
                                           # without this, read_stb() below can NEVER see
                                           # completion even after the instrument is done,
                                           # because ESR's OPC bit only reaches the Status
                                           # Byte summary (ESB) if it's enabled here first.
                self.lcr.write(scpi_exec)  # kick off the correction sweep
                self.lcr.write("*OPC")     # ask instrument to set OPC bit when done

                # IMPORTANT: while CORR:xxx:EXEC is still running, the
                # instrument's SCPI parser is busy and will NOT read a
                # follow-up query like "*ESR?" until it finishes the
                # command ahead of it — so sending queries to check
                # progress just queues up behind the still-running
                # operation and times out, no matter how many times you
                # retry. A GPIB serial poll (read_stb) is a separate,
                # lower-level bus transaction that the instrument's
                # interface answers immediately regardless of what the
                # parser is doing, so it's the correct way to check on
                # a long-running operation without blocking behind it.
                waited = 0.0
                done = False
                while waited < CAL_MAX_WAIT_S:
                    try:
                        stb = self.lcr.read_stb()
                        if stb & 0x20:   # ESB — Event Status Bit set
                            done = True
                            break
                    except pyvisa.errors.VisaIOError:
                        pass
                    time.sleep(POLL_INTERVAL_S)
                    waited += POLL_INTERVAL_S

                if not done:
                    raise TimeoutError(
                        f"{kind} calibration did not finish within "
                        f"{CAL_MAX_WAIT_S}s. Check that the GPIB cable "
                        f"is seated properly and the fixture is set up "
                        f"as instructed (fully open for OPEN cal, "
                        f"shorted for SHORT cal), then try again.")

                # Confirm no SCPI error was raised during the sweep.
                try:
                    self.lcr.timeout = 5000
                    err = self.lcr.query("SYST:ERR?").strip()
                    if not err.startswith("+0") and not err.startswith("0,"):
                        raise RuntimeError(
                            f"Instrument reported an error during "
                            f"{kind.lower()} calibration: {err}")
                except pyvisa.errors.VisaIOError:
                    pass  # some firmware revisions don't support SYST:ERR?

                self.lcr.write(scpi_enable)
                self.lcr.timeout = 10000
                self.root.after(0, self._on_cal_success,
                                kind, btn, label_base, done_attr)
            except Exception as e:
                try:
                    self.lcr.timeout = 10000
                except Exception:
                    pass
                self.root.after(0, self._on_cal_failure,
                                kind, btn, label_base, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_cal_success(self, kind, btn, label_base, done_attr):
        self.calibrating = False
        setattr(self, done_attr, True)
        self.save_calibration()
        btn.config(text=f"{label_base}  ✓", state="normal",
                  bg="#E8F8EE", fg=GREEN)
        self.log(f"{kind} calibration complete.")
        self._refresh_cal_status()

    def _on_cal_failure(self, kind, btn, label_base, error_msg):
        self.calibrating = False
        btn.config(text=f"{label_base}  ✗", state="normal",
                  bg="#FFF0F0", fg=RED)
        messagebox.showerror(f"{kind} Calibration Failed", error_msg)
        self.log(f"{kind} calibration ERROR: {error_msg}")
        self._refresh_cal_status()
    def save_calibration(self):
      data = {
        "open": self.open_cal_done,
        "short": self.short_cal_done
      }

      with open(self.calibration_file, "w") as f:
        json.dump(data, f)


    def load_calibration(self):
     if os.path.exists(self.calibration_file):
        with open(self.calibration_file, "r") as f:
            data = json.load(f)

        self.open_cal_done = data.get("open", False)
        self.short_cal_done = data.get("short", False)
        if self.open_cal_done:
            self.btn_open_cal.config(
                text="Open Cal ✓",
                bg=GREEN,
                fg="white"
            )

        if self.short_cal_done:
            self.btn_short_cal.config(
                text="Short Cal ✓",
                bg=GREEN,
                fg="white"
            )

    def _clear_calibration(self):
        if self.running:
            messagebox.showwarning(
                "Measurement Running",
                "Stop the current measurement before clearing calibration.")
            return
        self.open_cal_done  = False
        self.short_cal_done = False
        import os
        if os.path.exists(self.calibration_file):
         os.remove(self.calibration_file)
        self.btn_open_cal.config(text="Open Cal  ○", state="normal",
                                 bg="#EEF0FF", fg=VIOLET)
        self.btn_short_cal.config(text="Short Cal  ○", state="normal",
                                  bg="#EEF0FF", fg=VIOLET)
        if self.lcr is not None:
            try:
                self._drain_stale_response()
                self.lcr.write("CORR:OPEN OFF")
                self.lcr.write("CORR:SHORT OFF")
                self.lcr.write("*CLS")   # clear the error queue/status —
                                         # resets any stuck -420 Query
                                         # UNTERMINATED left over from a
                                         # previous failed attempt
            except Exception as e:
                self.log(f"Warning clearing instrument cal state: {e}")
        self.log("Calibration cleared.")
        self._refresh_cal_status()
        self.save_calibration()

    def _refresh_cal_status(self):
        """Keeps the warning label and Start button gate in sync
        with current open/short calibration state."""
        both_done = self.open_cal_done and self.short_cal_done
        if both_done:
            self.lbl_cal_status.config(
                text="✓  Open + Short calibration complete",
                fg=GREEN)
        elif self.open_cal_done or self.short_cal_done:
            missing = "Short" if self.open_cal_done else "Open"
            self.lbl_cal_status.config(
                text=f"⚠  {missing} calibration still required before Start",
                fg=AMBER)
        else:
            self.lbl_cal_status.config(
                text="⚠  Calibration required before Start is enabled",
                fg=AMBER)

        # Only touch Start's enabled state if we're not mid-measurement
        # (­_finish/_start manage that state during a run).
        if not self.running:
            self.btn_start.config(
                state="normal" if both_done else "disabled")

    # ──────────────────────────────────────────────────────────
    #  BROWSE CSV  (preserved)
    # ──────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"LCR_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        if path:
            self.csv_path.set(path)

    # ──────────────────────────────────────────────────────────
    #  EXPORT  (Phase 1)
    # ──────────────────────────────────────────────────────────
    def _export_clicked(self):
        if self.running:
            messagebox.showwarning(
                "Measurement Running",
                "Stop the current measurement before exporting.")
            return
        if not self.last_completed_run:
            messagebox.showinfo(
                "Nothing to Export",
                "Run a measurement to completion first — "
                "Export becomes available once readings have "
                "been collected and saved.")
            return

        # Small popup with the three export choices, styled to match
        # the rest of the app rather than a plain messagebox.
        win = tk.Toplevel(self.root)
        win.title("Export Results")
        win.configure(bg=BG_WHITE)
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text="📤  Export Results",
                bg=BG_WHITE, fg=T_DARK,
                font=F_SECTION).pack(padx=20, pady=(16, 4))
        tk.Label(win,
                text=f"{self.last_completed_run['meta'].get('n_readings', 0)} "
                     f"readings from this run",
                bg=BG_WHITE, fg=T_SECOND,
                font=F_SMALL).pack(padx=20, pady=(0, 12))

        btn_f = tk.Frame(win, bg=BG_WHITE)
        btn_f.pack(padx=20, pady=(0, 18))

        def choose(which):
            win.destroy()
            self._do_export(which)

        tk.Button(btn_f, text="📊  Excel (.xlsx)",
                 command=lambda: choose("excel"),
                 bg="#EEF0FF", fg=VIOLET, font=F_BTN,
                 relief="flat", cursor="hand2",
                 padx=14, pady=8, bd=0, width=16
        ).pack(pady=3)
        tk.Button(btn_f, text="📄  PDF Report",
                 command=lambda: choose("pdf"),
                 bg="#EEF0FF", fg=VIOLET, font=F_BTN,
                 relief="flat", cursor="hand2",
                 padx=14, pady=8, bd=0, width=16
        ).pack(pady=3)
        tk.Button(btn_f, text="📦  Both",
                 command=lambda: choose("both"),
                 bg=VIOLET, fg=T_WHITE, font=F_BTN,
                 relief="flat", cursor="hand2",
                 padx=14, pady=8, bd=0, width=16
        ).pack(pady=3)
        tk.Button(btn_f, text="Cancel",
                 command=win.destroy,
                 bg=BG_WHITE, fg=T_SECOND, font=F_SMALL,
                 relief="flat", cursor="hand2",
                 padx=10, pady=4, bd=0
        ).pack(pady=(8, 0))

    def _do_export(self, which):
        run = self.last_completed_run
        csv_path = run["csv_path"]
        meta = run["meta"]

        if not os.path.isfile(csv_path):
            messagebox.showerror(
                "File Missing",
                f"The CSV file for this run is no longer at:\n{csv_path}")
            return

        self.btn_export.config(state="disabled", text="Exporting...")
        self.root.update_idletasks()

        produced = []
        errors   = []

        # One-off optional title line, e.g. "DUC Capacitor (10 pF),
        # K = 0.1" — typed fresh each export, not persisted anywhere,
        # only used for the grouped Excel report's top row.
        duc_title = None
        temp_rh   = None
        if which in ("excel", "both"):
            duc_title = simpledialog.askstring(
                "DUC Info (optional)",
                "DUC value / K factor title for this report\n"
                "(e.g. \"DUC Capacitor (10 pF), K = 0.1\") — "
                "leave blank to skip:",
                parent=self.root)
            temp_rh = simpledialog.askstring(
                "Temp/RH (optional)",
                "Temperature / Humidity for this run\n"
                "(e.g. \"25.86°/52.99%\") — "
                "leave blank to fill in later in Excel:",
                parent=self.root)

        try:
            if which in ("excel", "both"):
                title = (f"{meta.get('modes','LCR')} Measurement — "
                         f"{meta.get('frequency','')}, "
                         f"{meta.get('voltage','')}, "
                         f"{meta.get('aperture','')} aperture, "
                         f"Average {meta.get('average','1')}")
                xlsx_path = export_csv_to_excel_grouped(
                    csv_path, title=title, duc_title=duc_title,
                    temp_rh=temp_rh or "")
                produced.append(xlsx_path)
        except Exception as e:
            errors.append(f"Excel export failed: {e}")

        try:
            if which in ("pdf", "both"):
                # Snapshot the current live graph as a PNG so the
                # PDF can embed it. Saved next to the CSV, named to
                # match, and cleaned up after the PDF is built.
                graph_png = os.path.splitext(csv_path)[0] + "_graph.png"
                try:
                    self.graph.fig.savefig(
                        graph_png, dpi=120,
                        facecolor=self.graph.fig.get_facecolor())
                except Exception:
                    graph_png = None  # PDF still generates without it

                pdf_path = export_csv_to_pdf(csv_path, meta, graph_png)
                produced.append(pdf_path)

                if graph_png and os.path.isfile(graph_png):
                    try:
                        os.remove(graph_png)
                    except OSError:
                        pass
        except Exception as e:
            errors.append(f"PDF export failed: {e}")

        self.btn_export.config(state="normal", text="Export")

        if produced:
            names = "\n".join(os.path.basename(p) for p in produced)
            self.log(f"Exported: {', '.join(os.path.basename(p) for p in produced)}")
            messagebox.showinfo(
                "Export Complete",
                f"Saved to the same folder as the CSV:\n\n{names}")
        if errors:
            messagebox.showerror("Export Error", "\n".join(errors))

    # ──────────────────────────────────────────────────────────
    #  START  (preserved)
    # ──────────────────────────────────────────────────────────
    def _start(self):
        if self.lcr is None:
            messagebox.showwarning("Not Connected",
                                   "Connect to instrument first.")
            return
        busy = self._instrument_busy()
        if busy:
            messagebox.showwarning(
                "Instrument Busy",
                f"Cannot start — {busy}. Stop it first.")
            return
        if not (self.open_cal_done and self.short_cal_done):
            messagebox.showwarning(
                "Calibration Required",
                "Run Open Cal and Short Cal before starting "
                "measurements.")
            return
        try:
            freq_hz = float(self.var_freq_hz.get())
            if not (20 <= freq_hz <= 2_000_000):
                raise ValueError
        except ValueError:
            messagebox.showerror("Frequency Error",
                                 "Frequency must be 20 Hz – 2 MHz.")
            return
        try:
            n        = int(self.var_n.get())
            interval = float(self.var_interval.get())
            if n < 1 or interval < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Input Error",
                                 "Readings ≥ 1, interval ≥ 0.")
            return

        active_modes = [MODES[i]
                        for i, v in enumerate(self.selected_modes)
                        if v.get()]
        if not active_modes:
            messagebox.showerror("Mode Error",
                                 "Select at least one mode.")
            return

        aperture = self.var_aperture.get()
        average  = self.var_average.get()
        if average not in AVERAGES:
            messagebox.showerror("Average Error",
                                 "Average must be one of "
                                 "1, 2, 4, 8, 16, 32, 64, 128, 256.")
            return
        volt     = self.var_volt_var.get()
        freq_str = (str(int(freq_hz))
                    if freq_hz == int(freq_hz) else str(freq_hz))

        # Stash this run's settings so _finish() can build export
        # metadata once the run completes (active_modes/aperture/volt
        # are local to _start(), not otherwise visible at finish time).
        self._current_run_meta = {
            "modes": ", ".join(active_modes),
            "frequency": _hz_label(freq_hz),
            "voltage": f"{volt} V",
            "aperture": aperture,
            "average": average,
            "instrument": self._idn_str.split(",")[1].strip()
                         if self._idn_str and "," in self._idn_str
                         else "Agilent E4980A",
        }

        csv_path = self.csv_path.get()
        if not csv_path:
            csv_path = f"LCR_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            self.csv_path.set(csv_path)
        try:
            self.csv_file   = open(csv_path, "w", newline="",
                                   encoding="utf-8-sig")
            self.csv_writer = csv.writer(self.csv_file)
            self.csv_writer.writerow([
                "Reading#","Timestamp","Elapsed(s)","Mode",
                "Freq(Hz)","Volt(V)","Aperture","Average",
                "P1_raw","P1_display","P2_raw","P2_display","Status"])
        except Exception as e:
            messagebox.showerror("CSV Error", str(e))
            return

        self._stats_data   = {m: [] for m in active_modes}
        self.reading_count = 0
        self._total_n      = n
        self.start_time    = time.time()
        self.running       = True
        self.paused        = False
        self._paused_seconds = 0.0
        self.pause_event.set()
        self.graph.clear()
        self._refresh_graph_mode_options(active_modes)

        self.btn_start.config(state="disabled")
        self.btn_pause.config(state="normal", text="⏸  PAUSE",
                              bg=AMBER, fg=T_DARK)
        self.btn_stop.config(state="normal")
        self.progressbar.config(value=0)
        self.lbl_freq_disp.config(text=_hz_label(freq_hz))
        self.lbl_volt_disp.config(text=f"{volt} V")
        self.lbl_aper_disp.config(text=aperture)
        self.lbl_status_pill.config(text="MEASURING", fg=AMBER)

        self.log(f"Start  {', '.join(active_modes)}  "
                 f"@ {_hz_label(freq_hz)}  {volt}V  {aperture}  "
                 f"×{n}  Δ{interval}s")

        threading.Thread(
            target=self._measure_loop,
            args=(active_modes, n, interval, aperture, freq_str, volt,
                  average),
            daemon=True).start()

    # ──────────────────────────────────────────────────────────
    #  STOP  (preserved + releases any pause so thread can exit)
    # ──────────────────────────────────────────────────────────
    def _stop(self):
        self.running = False
        self.paused  = False
        self.pause_event.set()   # release a paused thread so it can exit
        self.btn_pause.config(state="disabled", text="⏸  PAUSE",
                              bg=AMBER, fg=T_DARK)
        self.log("Stopped.")

    # ──────────────────────────────────────────────────────────
    #  FREQUENCY SWEEP  (Phase 2)
    #  Mirrors _start/_measure_loop's threading pattern: validate
    #  on the main thread, run the actual sweep on a background
    #  thread, marshal all GUI updates back via root.after(0, ...).
    # ──────────────────────────────────────────────────────────
    @staticmethod
    def _log_spaced_frequencies(f_start, f_stop, steps):
        """Generate `steps` frequencies log-spaced between f_start
        and f_stop (inclusive), without requiring numpy — this
        project has no other numpy dependency, so a plain log10/
        exponent interpolation keeps it dependency-free."""
        if steps < 2:
            return [f_start]
        log_start = math.log10(f_start)
        log_stop  = math.log10(f_stop)
        return [
            10 ** (log_start + (log_stop - log_start) * i / (steps - 1))
            for i in range(steps)
        ]

    def _start_sweep(self):
        if self.lcr is None:
            messagebox.showwarning("Not Connected",
                                   "Connect to instrument first.")
            return
        busy = self._instrument_busy()
        if busy:
            messagebox.showwarning(
                "Instrument Busy",
                f"Cannot start sweep — {busy}. Stop it first.")
            return

        try:
            f_start = float(self.var_sweep_start.get())
            f_stop  = float(self.var_sweep_stop.get())
            steps   = int(self.var_sweep_steps.get())
            if not (20 <= f_start <= 2_000_000):
                raise ValueError("Start frequency must be 20 Hz – 2 MHz.")
            if not (20 <= f_stop <= 2_000_000):
                raise ValueError("Stop frequency must be 20 Hz – 2 MHz.")
            if f_stop <= f_start:
                raise ValueError("Stop frequency must be greater than "
                                 "start frequency.")
            if steps < 2:
                raise ValueError("Steps must be at least 2.")
        except ValueError as e:
            messagebox.showerror("Sweep Input Error",
                                 str(e) if str(e) else
                                 "Enter valid numbers for frequency/steps.")
            return

        active_modes = [MODES[i]
                        for i, v in enumerate(self.selected_modes)
                        if v.get()]
        if not active_modes:
            messagebox.showerror("Mode Error",
                                 "Select at least one measurement "
                                 "mode before sweeping.")
            return

        aperture = self.var_aperture.get()
        average  = self.var_average.get()
        volt     = self.var_volt_var.get()

        base = self.csv_path.get() or \
            f"LCR_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        sweep_csv_path = os.path.splitext(base)[0] + "_sweep.csv"
        try:
            self.sweep_csv_file = open(sweep_csv_path, "w", newline="",
                                       encoding="utf-8-sig")
            self.sweep_csv_writer = csv.writer(self.sweep_csv_file)
            self.sweep_csv_writer.writerow([
                "Step", "Freq(Hz)", "Mode",
                "P1_raw", "P1_display", "P2_raw", "P2_display", "Status"])
        except Exception as e:
            messagebox.showerror("Sweep CSV Error", str(e))
            return

        self.sweep_data = {m: {"freqs": [], "p1": [], "p2": []}
                           for m in active_modes}
        self.sweep_graph.clear()
        self.sweep_graph.set_titles_for_modes(active_modes)

        self.sweeping = True
        self.btn_sweep_start.config(state="disabled")
        self.btn_sweep_stop.config(state="normal")
        self.sweep_progressbar.config(value=0)
        self.lbl_sweep_status.config(
            text=f"Sweeping {f_start:.0f} Hz → {f_stop:.0f} Hz "
                 f"in {steps} steps...", fg=AMBER)
        self.log(f"Sweep start  {', '.join(active_modes)}  "
                 f"{f_start:.0f}–{f_stop:.0f} Hz  ×{steps}")

        threading.Thread(
            target=self._sweep_loop,
            args=(active_modes, f_start, f_stop, steps, aperture, volt,
                  average),
            daemon=True).start()

    def _stop_sweep(self):
        self.sweeping = False
        self.btn_sweep_stop.config(state="disabled")
        self.log("Sweep stop requested.")

    def _sweep_loop(self, active_modes, f_start, f_stop, steps,
                    aperture, volt, average="1"):
        try:
            self.lcr.write(f"VOLT {volt}")
            # Same averaging factor as the main measurement loop —
            # changing Average must not require recalibration or
            # affect DC Bias, so nothing else is touched here.
            self.lcr.write(f"APER {aperture},{average}")
            # Full-precision FETCh? responses — see the note in
            # _measure_loop's setup for why this is needed (default
            # ASCII format truncates to 6 sig figs, below panel
            # precision).
            self.lcr.write("FORMat:ASCii:LONG ON")
            # Same root-cause fix as _measure_loop / _synchronized_fetch:
            # single-shot triggering instead of free-run INIT:CONT ON,
            # so every point is guaranteed measured at the frequency/
            # mode that was just set, not whatever was in the buffer.
            # NOTE: trigger source stays INT (not BUS) — with BUS,
            # INIT only *arms* the trigger system and waits for an
            # explicit *TRG that nothing here ever sends, so *OPC?
            # would block until timeout on every single reading. With
            # INT + INIT:CONT OFF, each INIT immediately fires one
            # measurement and *OPC? blocks only as long as that one
            # acquisition actually takes.
            self.lcr.write("TRIG:SOUR INT")
            self.lcr.write("INIT:CONT OFF")
            self.lcr.write("*CLS")
            self.lcr.query("*OPC?")
        except Exception as e:
            self.root.after(0, self.log, f"Sweep setup error: {e}")
            self.root.after(0, lambda err=e: messagebox.showerror(
                "Instrument Error",
                f"Failed to configure the instrument for sweep: {err}"))
            self.root.after(0, self._finish_sweep)
            return

        freqs = self._log_spaced_frequencies(f_start, f_stop, steps)

        for step, freq in enumerate(freqs, start=1):
            if not self.sweeping:
                break
            try:
                self.lcr.write(f"FREQ {freq:.6f}")
                # Frequency just changed even if the mode code didn't
                # -- force _synchronized_fetch to do a full settle
                # for every sweep point rather than skipping the
                # re-trigger (that skip is only safe when NOTHING
                # about the measurement setup has changed).
                self._current_synced_mode = None

                for mode in active_modes:
                    if not self.sweeping:
                        break
                    try:
                        p1_raw, p2_raw, st_raw = self._synchronized_fetch(
                            mode, aperture, average)

                        pt1, pt2 = MODE_PARAM_TYPES.get(
                            scpi_code(mode), ("R", "X"))
                        p1_disp = format_lcr_value(p1_raw, pt1, aperture)
                        p2_disp = format_lcr_value(p2_raw, pt2, aperture)
                        status  = self._decode_status(st_raw)

                        self.sweep_data[mode]["freqs"].append(freq)
                        self.sweep_data[mode]["p1"].append(p1_raw)
                        self.sweep_data[mode]["p2"].append(p2_raw)

                        self.sweep_csv_writer.writerow([
                            step, f"{freq:.3f}", mode,
                            p1_raw, ascii_safe(p1_disp),
                            p2_raw, ascii_safe(p2_disp), status])
                        self.sweep_csv_file.flush()

                    except Exception as e:
                        self.root.after(0, self.log,
                                        f"[Sweep {mode} @ {freq:.0f}Hz] {e}")
                        self._recover_after_fetch_error()

            except Exception as e:
                self.root.after(0, self.log,
                                f"[Sweep @ {freq:.0f}Hz] {e}")
                self._recover_after_fetch_error()

            pct = int(step * 100 / steps)
            self.root.after(0, self._update_sweep_display, pct, step, steps)

        # Hand the instrument back to free-run for the front panel.
        try:
            self.lcr.write("TRIG:SOUR INT")
            self.lcr.write("INIT:CONT ON")
        except Exception:
            pass

        self.root.after(0, self._finish_sweep)

    def _update_sweep_display(self, pct, step, total_steps):
        self.sweep_progressbar.config(value=pct)
        self.lbl_sweep_status.config(
            text=f"Step {step} / {total_steps}  ({pct}%)", fg=AMBER)
        self.sweep_graph.update_data(self.sweep_data)

    def _finish_sweep(self):
        self.sweeping = False
        self.btn_sweep_start.config(state="normal")
        self.btn_sweep_stop.config(state="disabled")
        if self.sweep_csv_file:
            self.sweep_csv_file.close()
            self.sweep_csv_file = None
        n_pts = sum(len(d["freqs"]) for d in self.sweep_data.values())
        self.lbl_sweep_status.config(
            text=f"Sweep complete — {n_pts} points collected", fg=GREEN)
        self.log(f"Sweep done — {n_pts} points collected.")

    # ──────────────────────────────────────────────────────────
    #  DC BIAS  (Phase 3)
    #  E4980A internal DC bias requires Option 001, range ±40 V.
    #  Uses the same background-thread + *OPC? confirmation
    #  pattern as Open/Short calibration, and is gated by the
    #  same _instrument_busy() check as every other GPIB operation.
    # ──────────────────────────────────────────────────────────
    def _validate_bias_voltage(self):
        try:
            v = float(self.var_bias_volt.get())
        except ValueError:
            messagebox.showerror("Bias Voltage Error",
                                 "Enter a numeric bias voltage.")
            return None
        if not (-40.0 <= v <= 40.0):
            messagebox.showerror("Bias Voltage Error",
                                 "Bias voltage must be between "
                                 "-40 V and +40 V (Option 001 range).")
            return None
        return v

    def _toggle_bias(self):
        if self.var_bias_enabled.get():
            self._enable_bias()
        else:
            self._disable_bias()

    def _enable_bias(self):
        if self.lcr is None:
            messagebox.showwarning("Not Connected",
                                   "Connect to the instrument first.")
            self.var_bias_enabled.set(False)
            return
        busy = self._instrument_busy()
        if busy:
            messagebox.showwarning(
                "Instrument Busy",
                f"Cannot enable bias — {busy}. Stop it first.")
            self.var_bias_enabled.set(False)
            return

        v = self._validate_bias_voltage()
        if v is None:
            self.var_bias_enabled.set(False)
            return

        self.bias_busy = True
        self.chk_bias_enable.config(state="disabled")
        self.lbl_bias_status.config(text="Applying...", fg=AMBER)

        def worker():
            try:
                self.lcr.timeout = 15000
                self.lcr.write(f"BIAS:VOLT {v}")
                self.lcr.write("BIAS:STAT ON")
                self.lcr.query("*OPC?")
                self.lcr.timeout = 10000
                self.root.after(0, self._on_bias_enabled, v)
            except Exception as e:
                self.root.after(0, self._on_bias_error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_bias_enabled(self, v):
        self.bias_busy = False
        self.bias_on   = True
        self.chk_bias_enable.config(state="normal")
        self.btn_bias_disable.config(state="normal")
        self.lbl_bias_status.config(text=f"ON  ({v:+.3f} V)", fg=GREEN)
        self.log(f"DC Bias enabled: {v:+.3f} V")

    def _on_bias_error(self, error_msg):
        self.bias_busy = False
        self.var_bias_enabled.set(False)
        self.chk_bias_enable.config(state="normal")
        self.lbl_bias_status.config(text="ERROR", fg=RED)
        messagebox.showerror("DC Bias Error", error_msg)
        self.log(f"DC Bias ERROR: {error_msg}")

    def _disable_bias(self):
        if self.lcr is None:
            self.var_bias_enabled.set(False)
            self.lbl_bias_status.config(text="OFF", fg=T_SECOND)
            return
        if self.bias_busy:
            return

        self.bias_busy = True
        self.btn_bias_disable.config(state="disabled")
        self.chk_bias_enable.config(state="disabled")

        def worker():
            try:
                self.lcr.write("BIAS:STAT OFF")
                self.lcr.query("*OPC?")
                self.root.after(0, self._on_bias_disabled)
            except Exception as e:
                self.root.after(0, self._on_bias_error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_bias_disabled(self):
        self.bias_busy = False
        self.bias_on   = False
        self.var_bias_enabled.set(False)
        self.chk_bias_enable.config(state="normal")
        self.btn_bias_disable.config(state="disabled")
        self.lbl_bias_status.config(text="OFF", fg=T_SECOND)
        self.log("DC Bias disabled.")

    # ──────────────────────────────────────────────────────────
    #  PAUSE / RESUME
    # ──────────────────────────────────────────────────────────
    def _toggle_pause(self):
        if not self.running:
            return
        if not self.paused:
            # → PAUSE: hold the measurement thread before its next cycle
            self.paused = True
            self.pause_event.clear()
            self._pause_started_at = time.time()
            self.btn_pause.config(text="▶  RESUME",
                                  bg=GREEN, fg=T_WHITE)
            self.lbl_status_pill.config(text="PAUSED", fg=ORANGE)
            self.log("Paused.")
        else:
            # → RESUME
            self.paused = False
            if self._pause_started_at is not None:
                self._paused_seconds += time.time() - self._pause_started_at
                self._pause_started_at = None
            self.pause_event.set()
            self.btn_pause.config(text="⏸  PAUSE",
                                  bg=AMBER, fg=T_DARK)
            self.lbl_status_pill.config(text="MEASURING", fg=AMBER)
            self.log("Resumed.")

    # ──────────────────────────────────────────────────────────
    #  MEASUREMENT LOOP  (100% preserved)
    # ──────────────────────────────────────────────────────────
    def _elapsed(self):
        """Elapsed time since start, with any paused duration subtracted
        out so pauses don't inflate the graph X-axis or CSV timestamps."""
        live_pause = (time.time() - self._pause_started_at
                     if self._pause_started_at is not None else 0.0)
        return (time.time() - self.start_time
               - self._paused_seconds - live_pause)

    # ──────────────────────────────────────────────────────────
    #  SYNCHRONIZED SINGLE-SHOT FETCH  (root-cause fix)
    #
    #  THE BUG: the previous code left the instrument in free-run
    #  continuous-trigger mode (TRIG:SOUR INT + INIT:CONT ON) and,
    #  after every FUNC:IMP mode change, just slept a fixed 0.25 s
    #  before calling FETCh?. FETCh? does not trigger a measurement
    #  — with INIT:CONT ON the instrument is triggering and
    #  overwriting its result buffer on its own schedule, and
    #  FETCh? simply returns whatever happens to be sitting in that
    #  buffer at the moment you call it. A fixed 0.25 s guess is
    #  only sometimes long enough:
    #    - SHORT aperture / low averaging: one acquisition is a few
    #      ms, so 0.25 s comfortably covers it → readings often
    #      matched (explains "larger capacitors sometimes match
    #      perfectly", issue #6, when those tests happened to use
    #      fast settings).
    #    - MED/LONG aperture with higher averaging (exactly what you
    #      need for a stable reading on 1-10 pF, issue #5/#13): one
    #      acquisition can take from tens of ms up to multiple
    #      seconds (aperture time × averaging factor). 0.25 s isn't
    #      close, so FETCh? returns a result from *before* the
    #      FUNC:IMP/FREQ/VOLT/APER change — i.e. a stale reading
    #      computed under the previous configuration. Because the
    #      instrument recomputes both displayed parameters from the
    #      same raw Z/θ measurement, and ranging can shift mid-way,
    #      it's common for only one of the two (e.g. P1) to have
    #      caught up by the time you read — explaining issue #2.
    #      This single mechanism is the root cause of issues #1-#8.
    #
    #  THE FIX: put the instrument in single-shot triggered mode
    #  (TRIG:SOUR INT + INIT:CONT OFF) once per run, then for every
    #  reading: write FUNC:IMP, issue INIT (which — with INT as the
    #  trigger source — fires that one measurement immediately, no
    #  separate trigger command needed), and block on *OPC? before
    #  calling FETCh?. Per SCPI, *OPC? cannot answer until the
    #  instrument has actually finished the pending operation — so
    #  the wait time automatically stretches to cover whatever
    #  aperture/averaging/auto-ranging/settling that specific
    #  reading needs, instead of a single guessed constant. FETCh?
    #  immediately after is then guaranteed to return the result of
    #  that exact INIT, under the settings that were active when it
    #  was issued — never a stale buffer.
    #  (An earlier revision of this fix used TRIG:SOUR BUS, which is
    #  wrong: with BUS, INIT only *arms* the trigger system and
    #  waits for an explicit *TRG that nothing here sends, so every
    #  *OPC? would block until timeout and no reading would ever
    #  come back — INT is the correct source for this pattern.)
    # ──────────────────────────────────────────────────────────
    @staticmethod
    def _measurement_timeout_ms(aperture, average):
        """Conservative upper bound on how long a single *OPC?-gated
        acquisition may take, so the PyVISA read timeout never fires
        before the instrument legitimately finishes (which the old
        fixed-sleep code never risked, since it didn't actually wait
        on completion). These per-aperture base times are a safety
        margin, not a datasheet figure — if you see VI_ERROR_TMO on
        LONG aperture with high averaging, raise the base numbers
        here; it will never cause a wrong *reading*, only a slower
        failure mode than before."""
        base_ms = {"SHORT": 50, "MED": 200, "LONG": 800}.get(
            aperture.strip().upper(), 200)
        try:
            avg = max(1, int(average))
        except (TypeError, ValueError):
            avg = 1
        return max(10000, base_ms * avg * 3 + 3000)

    def _synchronized_fetch(self, mode, aperture, average):
        """Switch to `mode` and return a (p1_raw, p2_raw, status_raw)
        tuple for it. Raises on any SCPI/parse failure — caller is
        responsible for recovery (see _drain_stale_response usage at
        each call site).

        EXACT-MATCH MODE (changed 2026-07-16): every call now forces
        a fresh single-shot acquisition (FUNC:IMP if needed + INIT:IMM
        + *OPC?) before FETCh?, instead of only doing this on a mode
        change and free-running afterward. The free-run approach
        mirrored the panel's own display loop, but it meant the
        software's read and the panel's on-screen value were two
        independent taps into the same live buffer — close, but a
        sample apart, which showed up as a few counts of drift in the
        last 1-2 digits on jitter-sensitive parameters (L, C, Rs)
        even though D/Q (less sensitive to that jitter) matched
        exactly. Forcing a real trigger+OPC-gated measurement on every
        read makes each value the freshest possible completed
        measurement, closest to what the panel is showing at read
        time. Trade-off: each reading now takes one full aperture
        cycle instead of a near-instant buffer peek, so throughput is
        lower — this matters most at LONG aperture with many
        readings/modes selected.
        """
        code = MODE_SCPI[mode]
        old_timeout = self.lcr.timeout
        self.lcr.timeout = self._measurement_timeout_ms(aperture, average)
        mode_changed = (code != getattr(self, "_current_synced_mode", None))
        try:
            t0 = time.time()
            if mode_changed:
                print("Sending Mode =", code)
                self.lcr.write(f"FUNC:IMP {code}")
                print("Instrument Mode =", self.lcr.query("FUNC:IMP?"))
                self._current_synced_mode = code

            # Forced, single-shot, OPC-gated acquisition on EVERY
            # call now (not just on mode change) — see EXACT-MATCH
            # MODE note above.
            self.lcr.write("TRIG:SOUR INT")
            self.lcr.write("INIT:CONT OFF")
            self.lcr.write("INIT:IMM")
            print(f"  INIT:IMM sent (+{time.time()-t0:.3f}s)")
            self.lcr.query("*OPC?")   # blocks until this one
                                      # measurement has actually
                                      # completed
            print(f"  *OPC? returned (+{time.time()-t0:.3f}s)")

            raw = self.lcr.query("FETCh?").strip()
            print(f"  FETCh? returned (+{time.time()-t0:.3f}s)  RAW =", raw)
        except Exception as e:
            # Surface *exactly* where in the sequence it died, and
            # what the instrument itself thinks went wrong — this is
            # the piece that's been invisible so far (the "Sending
            # Mode" loop repeating with no RAW/P1/P2 ever printing).
            print(f"  ‼ fetch failed after {time.time()-t0:.3f}s: {e}")
            try:
                err = self.lcr.query("SYST:ERR?").strip()
                print("  SYST:ERR? =", err)
            except Exception as e2:
                print("  (SYST:ERR? query also failed:", e2, ")")
            raise
        finally:
            self.lcr.timeout = old_timeout

        parts = raw.split(",")
        if len(parts) < 2:
            raise ValueError(f"Unexpected FETCh? response: {raw!r}")
        p1_raw = float(parts[0])
        p2_raw = float(parts[1])
        st_raw = parts[2].strip() if len(parts) > 2 else "0"
        print("P1 =", p1_raw)
        print("P2 =", p2_raw)
        return p1_raw, p2_raw, st_raw

    def _recover_after_fetch_error(self):
        """A malformed/partial FETCh? response (parse failure,
        VI_ERROR_TMO mid-response, etc.) can leave unread bytes on
        the bus, which desyncs every subsequent query in the run —
        the classic symptom of that is 'Parameter 1 correct,
        Parameter 2 wrong' on the *next* reading, not the one that
        actually errored (issue #2). The old code just logged the
        exception and moved on, leaving that desync in place. This
        clears it before the next mode/cycle continues."""
        try:
            self.lcr.clear()
        except Exception:
            pass
        self._drain_stale_response()
        try:
            self.lcr.write("*CLS")
        except Exception:
            pass

    def _measure_loop(self, active_modes, n, interval,
                      aperture, freq_str, volt, average="1"):
        # Force the first _synchronized_fetch call of this run to do
        # a full settle, regardless of what mode a previous run left
        # the buffer synced to.
        self._current_synced_mode = None
        try:
            self.lcr.write(f"FREQ {freq_str}")
            self.lcr.write(f"VOLT {volt}")
            # Averaging factor is sent as the second APER parameter
            # (valid range on the E4980A is 1-256, matching AVERAGES).
            self.lcr.write(f"APER {aperture},{average}")
            # ROOT CAUSE FIX (2026-07-16): the E4980A's FETCh? response
            # defaults to a SHORT ASCII format (SN.NNNNNESNN — only 6
            # significant digits), which is *less* precise than the
            # front-panel display. This was confirmed directly from a
            # console capture: RAW = +9.98113E-12,... (6 sig figs) vs
            # the panel showing 9.981131 pF (7 sig figs) for the same
            # reading. The formatter was correctly reproducing every
            # digit it was given — there just weren't enough digits in
            # what the instrument sent. :FORMat:ASCii:LONG ON switches
            # FETCh?/MEAS? results to the long ASCII format (per the
            # E4980A/AL programming manual, this yields results like
            # +1.059517689E-24 — 9-10 significant digits), which is
            # enough to match or exceed the panel's own precision.
            self.lcr.write("FORMat:ASCii:LONG ON")
            # Single-shot synchronous acquisition, not free-run — see
            # _synchronized_fetch above for why this matters. Trigger
            # source stays INT (not BUS): with BUS, INIT only arms
            # the trigger and waits for an explicit *TRG that's never
            # sent, so every *OPC? below would block until timeout
            # and no reading would ever come back. With INT +
            # INIT:CONT OFF, each INIT fires one measurement
            # immediately and *OPC? blocks only for that acquisition.
            self.lcr.write("TRIG:SOUR INT")
            self.lcr.write("INIT:CONT OFF")
            self.lcr.write("*CLS")
            self.lcr.query("*OPC?")   # confirm FREQ/VOLT/APER have
                                      # actually been accepted/settled
                                      # before the very first reading
            # DIAGNOSTIC: read back what the instrument actually
            # latched for trigger source / continuous-init, since
            # *OPC? blocking forever on every reading (the symptom
            # you're seeing — "Sending Mode" repeats but "RAW ="
            # never prints) is the exact behavior you'd get if either
            # of these didn't actually take effect (e.g. front panel
            # still in a HOLD/local state overriding it, or this
            # firmware wanting the full "INITiate:CONTinuous OFF"
            # spelled out).
            print("TRIG:SOUR readback =", self.lcr.query("TRIG:SOUR?").strip())
            print("INIT:CONT readback =", self.lcr.query("INIT:CONT?").strip())
            print("FORMat:ASCii:LONG readback =",
                  self.lcr.query("FORMat:ASCii:LONG?").strip())
            print("SYST:ERR? after setup =", self.lcr.query("SYST:ERR?").strip())
        except Exception as e:
            self.root.after(0, self.log, f"Setup error: {e}")
            self.root.after(0, lambda err=e: messagebox.showerror(
                "Instrument Error",
                f"Failed to configure the instrument (Frequency/"
                f"Voltage/Aperture/Average): {err}"))
            self.root.after(0, self._finish)
            return

        for cycle in range(n):
            if not self.running:
                break
            cycle_results = {}

            for mode in active_modes:
                if not self.running:
                    break
                try:
                    p1_raw, p2_raw, st_raw = self._synchronized_fetch(
                        mode, aperture, average)

                    pt1, pt2 = MODE_PARAM_TYPES.get(scpi_code(mode), ("R","X"))
                    p1_disp  = format_lcr_value(p1_raw, pt1, aperture)
                    p2_disp  = format_lcr_value(p2_raw, pt2, aperture)
                    status   = self._decode_status(st_raw)

                    cycle_results[mode] = (p1_raw, p1_disp,
                                           p2_raw, p2_disp, status)
                    self._stats_data[mode].append(p1_raw)

                    self.reading_count += 1
                    elapsed = self._elapsed()
                    self.graph.add_point(mode, p1_raw, p2_raw, elapsed)

                    ts = datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S.%f")[:-3]
                    self.csv_writer.writerow([
                        cycle + 1, ts, f"{elapsed:.3f}",
                        mode, freq_str, volt, aperture, average,
                        p1_raw, ascii_safe(p1_disp),
                        p2_raw, ascii_safe(p2_disp), status])
                    self.csv_file.flush()

                    self.readings_log.append({
                        "n": self.reading_count,
                        "time": ts.split(" ")[-1][:8],
                        "mode": mode,
                        "p1": p1_disp,
                        "p2": p2_disp,
                        "status": status,
                    })

                except Exception as e:
                    self.root.after(0, self.log, f"[{mode}] {e}")
                    self._recover_after_fetch_error()

            elapsed = self._elapsed()
            self.root.after(0, self._update_display,
                            cycle_results, active_modes,
                            aperture, elapsed, n, cycle + 1)

            # ── PAUSE HOLD POINT ───────────────────────────────
            # Current full cycle has just finished — if Pause was
            # pressed, block here (no GPIB calls, no CPU spin) until
            # Resume sets the event again, or Stop releases it.
            self.pause_event.wait()
            if not self.running:
                break

            time.sleep(interval)

        # Hand the instrument back to free-run so the front panel
        # keeps live-updating for anyone standing at the bench after
        # the automated run ends.
        try:
            self.lcr.write("TRIG:SOUR INT")
            self.lcr.write("INIT:CONT ON")
        except Exception:
            pass

        self.root.after(0, self._finish)

    # ──────────────────────────────────────────────────────────
    #  UPDATE DISPLAY  (preserved + enhanced)
    # ──────────────────────────────────────────────────────────
    def _autosize_value_font(self, label, text, is_primary_row):
        """Shrink the big live-reading font as the string gets longer,
        so an unusually long value (many digits, a long unit suffix,
        etc.) shrinks to fit instead of visually overflowing the
        panel and getting clipped at the window edge."""
        base_size = 38 if is_primary_row else 22
        length = len(text or "")
        if length > 12:
            size = max(14, base_size - (length - 12) * 2)
        else:
            size = base_size
        label.config(font=(MONO, size, "bold"))

    def _update_display(self, results, active_modes,
                        aperture, elapsed, total_n, cycle_num=None):
        # Counter + progress
        # NOTE: reading_count increments once per active mode per
        # cycle, so with >1 mode selected it runs ahead of the
        # requested number of readings (e.g. 10 sub-readings for
        # n=5 with 2 modes active). The visible "x / n" counter must
        # track completed cycles instead, or it overshoots the
        # requested total (previously showed e.g. "10 / 5").
        if cycle_num is None:
            cycle_num = self.reading_count
        pct = int(self.reading_count * 100 /
                  max(total_n * len(active_modes), 1))
        self.lbl_n_disp.config(
            text=f"{cycle_num} / {total_n}")
        self.lbl_prog.config(text=f"{pct}%")
        self.progressbar.config(value=pct)

        # Elapsed HH:MM:SS
        h = int(elapsed) // 3600
        m = (int(elapsed) % 3600) // 60
        s = int(elapsed) % 60
        self.lbl_elapsed.config(text=f"{h:02d}:{m:02d}:{s:02d}")

        # Mode reading rows
        for i in range(3):
            lm, lp1n, lp1v, lp2n, lp2v, lst = self._mode_rows[i]
            if i < len(active_modes):
                mode = active_modes[i]
                if mode in results:
                    try:
                        p1r, p1d, p2r, p2d, st = results[mode]
                        pt1, pt2 = MODE_PARAM_TYPES.get(
                            scpi_code(mode), ("?","?"))

                        # Split unit from value for large display
                        p1_parts = p1d.rsplit(" ", 1)
                        p1_num   = p1_parts[0] if len(p1_parts) > 1 else p1d
                        p1_unit  = p1_parts[1] if len(p1_parts) > 1 else ""

                        lm.config(text=mode)
                        lp1n.config(text=pt1)
                        lp1v.config(text=p1_num)
                        lp2n.config(text=pt2)
                        lp2v.config(text=p2d)
                        self._autosize_value_font(lp1v, p1_num, i == 0)
                        self._autosize_value_font(lp2v, p2d, i == 0)
                        lst.config(
                            text=f"  {st}  ",
                            bg="#E8FFF0" if st=="OK" else "#FFE8E8",
                            fg=GREEN if st=="OK" else RED)
                        # Update unit label
                        if i < len(self._reading_p1_labels):
                            self._reading_p1_labels[i][1].config(
                                text=p1_unit)
                        self.lbl_status_pill.config(
                            text=st,
                            fg=GREEN if st=="OK" else RED)
                    except Exception as e:
                        # A formatting/display bug for THIS mode's
                        # value must not freeze the whole panel (this
                        # is exactly what produced the "counter moves,
                        # values stuck at — — —" symptom before — the
                        # exception used to propagate up through
                        # root.after and vanish silently in a windowed
                        # EXE). Log the real cause and show it plainly
                        # instead of leaving stale dashes on screen.
                        self.log(f"⚠ Display error [{mode}]: {e}  "
                                 f"(raw: {results.get(mode)})")
                        try:
                            log_path = os.path.join(
                                os.path.dirname(os.path.abspath(sys.argv[0])),
                                "LCR_error_log.txt")
                            with open(log_path, "a", encoding="utf-8") as f:
                                f.write(
                                    f"\n[{datetime.now()}] _update_display "
                                    f"mode={mode} results={results.get(mode)}\n"
                                    + traceback.format_exc() + "\n")
                        except Exception:
                            pass
                        lst.config(text="  ERR  ", bg="#FFE8E8", fg=RED)
            else:
                for lb in [lm, lp1n, lp1v, lp2n, lp2v]:
                    lb.config(text="")
                lst.config(text="", bg=BG_WHITE)

        # Stats
        for i in range(3):
            lm, lmin, lmax, lmean, lstd = self._stat_rows[i]
            if i < len(active_modes):
                mode = active_modes[i]
                vals = self._stats_data.get(mode, [])
                pt1, _ = MODE_PARAM_TYPES.get(scpi_code(mode), ("R","X"))
                lm.config(text=mode)
                if vals:
                    lmin.config(
                        text=f"min  {format_lcr_value(min(vals),pt1,aperture)}")
                    lmax.config(
                        text=f"max  {format_lcr_value(max(vals),pt1,aperture)}")
                    lmean.config(
                        text=f"avg  {format_lcr_value(sum(vals)/len(vals),pt1,aperture)}")
                    if len(vals) > 1:
                        lstd.config(
                            text=f"std  {format_lcr_value(_stats.stdev(vals),pt1,aperture)}")
            else:
                for lb in [lm, lmin, lmax, lmean, lstd]:
                    lb.config(text="")

        # Legend / caption — shows which mode is currently graphed
        graphed = self.graph.selected_mode or (
            active_modes[0] if active_modes else "")
        self.lbl_legend.config(
            text=f"Showing: {graphed}   — P1 (cyan)   — P2 (pink)"
            if graphed else "— P1 (cyan)   — P2 (pink)")

        self.graph.redraw(active_modes, aperture)
        self._refresh_datalogger(results, active_modes, aperture)

    # ──────────────────────────────────────────────────────────
    #  FINISH  (preserved)
    # ──────────────────────────────────────────────────────────
    def _finish(self):
        self.running = False
        self.paused  = False
        self.pause_event.set()
        self.btn_start.config(
            state="normal" if (self.open_cal_done and self.short_cal_done)
            else "disabled")
        self.btn_pause.config(state="disabled", text="⏸  PAUSE",
                              bg=AMBER, fg=T_DARK)
        self.btn_stop.config(state="disabled")
        self.lbl_status_pill.config(text="STABLE", fg=GREEN)
        if self.csv_file:
            self.csv_file.close()
            self.csv_file = None

        self._record_session()
        self._save_session_to_disk()

        # Make this run's CSV + metadata available for export, only
        # once the run has actually finished writing.
        if self.reading_count > 0:
            meta = dict(getattr(self, "_current_run_meta", {}))
            meta["n_readings"] = self.reading_count
            self.last_completed_run = {
                "csv_path": self.csv_path.get(),
                "meta": meta,
            }
            self.lbl_export_hint.config(
                text=f"{self.reading_count} readings ready", fg=GREEN)

        self.log(
            f"Done — {self.reading_count} readings → "
            f"{self.csv_path.get()}")

    # ──────────────────────────────────────────────────────────
    #  HELPERS  (preserved)
    # ──────────────────────────────────────────────────────────
    @staticmethod
    def _decode_status(raw):
        try:
            c = int(float(raw))
        except Exception:
            return raw
        return {0:"OK",1:"OVERLOAD",2:"NO DATA",
                4:"AVERAGING"}.get(c, f"CODE {c}")

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt_log.config(state="normal")
        self.txt_log.insert("end", f"[{ts}]  {msg}\n")
        self.txt_log.see("end")
        self.txt_log.config(state="disabled")

    def _draw_grad_bar(self, c):
        pass   # called on resize — canvas redraws via bind

    def _handle_tk_callback_exception(self, exc_type, exc_value, exc_tb):
        """Installed as root.report_callback_exception. Tkinter calls
        this instead of raising when a bound callback (including every
        root.after(0, ...) used to push live readings to the GUI)
        throws. Without this override the exception is printed to
        stderr and swallowed — invisible in a windowed/--noconsole
        EXE — and the run silently stalls right where it failed.
        Here we surface it both in the on-screen log and in a plain
        text file next to the script/EXE, and make sure the run
        doesn't just hang forever afterwards."""
        msg = "".join(traceback.format_exception(
            exc_type, exc_value, exc_tb))
        try:
            log_path = os.path.join(
                os.path.dirname(os.path.abspath(sys.argv[0])),
                "LCR_error_log.txt")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n[{datetime.now()}]\n{msg}\n")
        except Exception:
            pass
        try:
            self.log(f"⚠ Internal error: {exc_value}  "
                     f"(full trace in LCR_error_log.txt)")
        except Exception:
            pass
        # A callback exception during a run most often means a GUI
        # update failed mid-cycle — stop cleanly rather than leaving
        # the panel frozen with Start/Pause/Stop in an inconsistent
        # state.
        if self.running:
            try:
                self._stop()
            except Exception:
                pass


# ════════════════════════════════════════════════════════════
#  STANDALONE CLI TOOL  (merged in from Automation.py)
#  Run with:  python LCR_Automation.py --cli
#  Terminal-based, no GUI — separate from LCRApp above.
#  Uses CLI_MODES / CLI_* names so nothing collides with the
#  GUI's own MODES list or state.
# ════════════════════════════════════════════════════════════
CLI_MODES = {
    "CPD": ("Cp", "D"), "CPQ": ("Cp", "Q"), "CPG": ("Cp", "G"), "CPRP": ("Cp", "Rp"),
    "CSD": ("Cs", "D"), "CSQ": ("Cs", "Q"), "CSRS": ("Cs", "Rs"),
    "LPD": ("Lp", "D"), "LPQ": ("Lp", "Q"), "LPG": ("Lp", "G"),
    "LPRP": ("Lp", "Rp"), "LPRDC": ("Lp", "Rdc"),
    "LSD": ("Ls", "D"), "LSQ": ("Ls", "Q"), "LSRS": ("Ls", "Rs"),
    "ZTD": ("Z", "Theta"), "ZTR": ("Z", "R"),
    "YTD": ("Y", "Theta"), "YTR": ("Y", "R"),
    "RX": ("R", "X"),
    "GB": ("G", "B"),
    "VDCIDC": ("Vdc", "Idc"),
}


def run_cli_automation():
    import csv as _csv

    cli_rm = pyvisa.ResourceManager()
    cli_lcr = cli_rm.open_resource("GPIB0::17::INSTR")

    print("\nConnected Instrument:")
    print(cli_lcr.query("*IDN?"))

    print("\nAvailable Modes:\n")
    for mode_name in CLI_MODES:
        print(mode_name)

    mode = input("\nEnter Mode : ").upper()
    if mode not in CLI_MODES:
        print("Invalid Mode")
        return

    freq_input = input(
        "\nEnter Frequency (Example: 100Hz, 1KHz, 10KHz, 100KHz, 1MHz): "
    )
    freq_input = freq_input.upper().replace(" ", "")

    if "MHZ" in freq_input:
        frequency = float(freq_input.replace("MHZ", "")) * 1000000
    elif "KHZ" in freq_input:
        frequency = float(freq_input.replace("KHZ", "")) * 1000
    elif "HZ" in freq_input:
        frequency = float(freq_input.replace("HZ", ""))
    else:
        frequency = float(freq_input)
    frequency = int(frequency)

    number_of_readings = int(input("\nNumber of Readings : "))
    sampling_interval = float(input("Sampling Interval (sec): "))

    cli_lcr.write(f"FUNC:IMP {scpi_code(mode)}")
    cli_lcr.write(f"FREQ {frequency}")

    filename = "Measurement_Data.csv"
    file = open(filename, "w", newline='')
    writer = _csv.writer(file)

    p1_name, p2_name = CLI_MODES[mode]

    writer.writerow([
        "Reading No", "Timestamp", "Elapsed Time(s)", "Mode",
        "Frequency(Hz)", p1_name, p2_name, "Status"
    ])

    plt.ion()
    fig, axs = plt.subplots(2, 1, figsize=(8, 8))
    plt.subplots_adjust(hspace=0.5)

    time_data, p1_data, p2_data = [], [], []
    line1, = axs[0].plot([], [])
    line2, = axs[1].plot([], [])
    axs[0].set_title(f"{p1_name} vs Time")
    axs[1].set_title(f"{p2_name} vs Time")
    axs[0].grid()
    axs[1].grid()

    start_time = time.time()

    for i in range(number_of_readings):
        timestamp = datetime.now()
        elapsed_time = round(time.time() - start_time, 2)

        data = cli_lcr.query("FETCH?")
        parts = data.strip().split(",")
        p1 = float(parts[0])
        p2 = float(parts[1])
        status = parts[2]

        writer.writerow([
            i + 1, timestamp, elapsed_time, mode,
            frequency, p1, p2, status
        ])
        file.flush()

        print("\n" + "=" * 50)
        print("Reading", i + 1)
        print("Time =", elapsed_time)
        print(f"{p1_name} = {p1}")
        print(f"{p2_name} = {p2}")
        print("Status =", status)

        time_data.append(elapsed_time)
        p1_data.append(p1)
        p2_data.append(p2)
        line1.set_data(time_data, p1_data)
        line2.set_data(time_data, p2_data)
        axs[0].relim(); axs[0].autoscale_view()
        axs[1].relim(); axs[1].autoscale_view()
        plt.pause(0.05)

        time.sleep(sampling_interval)

    file.close()
    cli_lcr.close()
    cli_rm.close()
    plt.ioff()
    plt.show()

    print("\nMeasurement Completed")
    print("CSV file saved as Measurement_Data.csv")


# ════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    if "--cli" in sys.argv:
        run_cli_automation()
        raise SystemExit

    # ── Windows DPI awareness ───────────────────────────────
    # Without this, on any PC where Windows display scaling is
    # not 100% (125%/150% are common on laptops), Windows just
    # bitmap-scales the whole Tk window after the fact instead
    # of letting it render at its true pixel size. That silently
    # eats into the fixed-width bottom toolbar row and was why
    # "🔬 Fit to Variation" (the last button, packed side="right")
    # would get clipped off-window on some machines but not
    # others — nothing to do with the zoom code itself.
    if sys.platform == "win32":
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass

    root = tk.Tk()

    style = ttk.Style(root)
    style.theme_use("clam")

    # Combobox styling — white/light theme for settings card
    style.configure("TCombobox",
        fieldbackground="#EEF0FF",
        background="#EEF0FF",
        foreground=T_DARK,
        selectbackground="#5B5FFF",
        selectforeground=T_WHITE,
        arrowcolor=VIOLET,
        bordercolor="#D0D0E8",
        lightcolor="#EEF0FF",
        darkcolor="#EEF0FF",
        font=F_BODY)
    style.map("TCombobox",
        fieldbackground=[("readonly", "#EEF0FF")],
        foreground=[("readonly", T_DARK)],
        bordercolor=[("focus", VIOLET)])

    LCRApp(root)
    root.mainloop()
